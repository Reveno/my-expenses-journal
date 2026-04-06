import asyncio, io, json, logging, re, os, time, random, string, sqlite3, urllib.request
try:
    import psycopg2
    import psycopg2.extras
    PSYCOPG2_OK = True
except ImportError as _pg_err:
    PSYCOPG2_OK = False
    class _FakePg:
        def __getattr__(self, name):
            raise ImportError(
                f"psycopg2 is required when DATABASE_URL is set. "
                f"Run: pip install psycopg2-binary"
            )
    psycopg2 = _FakePg()
from datetime import datetime, timedelta
from collections import defaultdict
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ContextTypes, filters, ConversationHandler
)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ── Завантаження .env (тільки локально, на Railway не потрібно) ──────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv не встановлено — ок на Railway

TOKEN   = os.getenv("BOT_TOKEN", "")
if not TOKEN:
    print("\n" + "="*55)
    print("  ПОМИЛКА: BOT_TOKEN не знайдено!")
    print("  1. Встанови залежності: pip install -r requirements.txt")
    print("  2. Переконайся що файл .env є в папці з bot.py")
    print("  3. У .env має бути рядок: BOT_TOKEN=твій_токен")
    print("="*55 + "\n")
    raise SystemExit(1)
DATABASE_URL = os.getenv("DATABASE_URL", "")  # set by Railway PostgreSQL plugin
DB_PATH      = os.getenv("DB_PATH", "expenses.db")  # used when DATABASE_URL is empty
USE_PG       = bool(DATABASE_URL)  # True on Railway, False locally
DONATE_URL = os.getenv("DONATE_URL", "")   # ваша ссылка PayPal / Mono / Ko-fi
_raw_admin = os.getenv("ADMIN_ID", "").strip()
ADMIN_ID   = int(_raw_admin) if _raw_admin.isdigit() else 0  # ваш Telegram ID для отримання фідбеку
_feedback_cooldown: dict[int, float] = {}  # uid → timestamp останнього фідбеку
MAX_INPUT  = 100
MAX_AMOUNT = 999_999.99
_raw_allow = os.getenv("ALLOWED_USERS", "")
ALLOWED_USERS: frozenset[int] = frozenset(
    int(x) for x in _raw_allow.split(",") if x.strip().isdigit()
)

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)

# ── Стани ──────────────────────────────────────────────────────────────────────
(CHOOSE_CATEGORY, ENTER_AMOUNT, ENTER_NAME,
 SET_LIMIT_CAT, SET_LIMIT_AMOUNT,
 TMPL_ACTION, TMPL_ADD_NAME, TMPL_ADD_AMOUNT, TMPL_ADD_CAT, TMPL_DEL,
 LANG_SELECT, CURR_PRIMARY, CURR_SECONDARY,
 CONVERT,
 EXPORT_PERIOD,
 CCAT_MENU, CCAT_NAME, CCAT_DEL,
 RECUR_MENU, RECUR_NAME, RECUR_AMT, RECUR_CAT, RECUR_DAY, RECUR_DEL,
 REMIND_MENU, REMIND_INACT, REMIND_DAILY, REMIND_WDAY, REMIND_WTIME,
 SETTINGS_MENU,
 ONBOARD_LANG, ONBOARD_CUR_PRI, ONBOARD_CUR_SEC,
 FEEDBACK_MSG,
 INCOME_AMT, INCOME_SRC,
 FINANCE_MENU, REPORTS_MENU, MORE_MENU) = range(39)

# ── Валюти ─────────────────────────────────────────────────────────────────────
CURRENCY_SYMBOLS = {"UAH": "₴", "USD": "$", "EUR": "€", "GBP": "£", "PLN": "zł"}
CURRENCY_BUTTONS = {
    "UAH": "₴ UAH — Гривня / Hryvnia",
    "USD": "$ USD — Долар / Dollar",
    "EUR": "€ EUR — Євро / Euro",
}
LANG_BUTTONS = {
    "uk": "🇺🇦 Українська",
    "ru": "🇷🇺 Русский",
    "en": "🇬🇧 English",
    "de": "🇩🇪 Deutsch",
}

# ── Вбудовані категорії ────────────────────────────────────────────────────────
BUILT_IN_KEYS = ["food","transport","entertainment","housing",
                 "clothes","health","education","pets","other"]
CATEGORY_LABELS = {
    "ru": {"food":"🍔 Еда","transport":"🚌 Транспорт","entertainment":"🎮 Развлечения",
           "housing":"🏠 Жильё","clothes":"👕 Одежда","health":"💊 Здоровье",
           "education":"📚 Образование","pets":"🐾 Питомцы","other":"🔧 Другое"},
    "uk": {"food":"🍔 Їжа","transport":"🚌 Транспорт","entertainment":"🎮 Розваги",
           "housing":"🏠 Житло","clothes":"👕 Одяг","health":"💊 Здоров'я",
           "education":"📚 Освіта","pets":"🐾 Тварини","other":"🔧 Інше"},
    "en": {"food":"🍔 Food","transport":"🚌 Transport","entertainment":"🎮 Entertainment",
           "housing":"🏠 Housing","clothes":"👕 Clothes","health":"💊 Health",
           "education":"📚 Education","pets":"🐾 Pets","other":"🔧 Other"},
    "de": {"food":"🍔 Essen","transport":"🚌 Transport","entertainment":"🎮 Freizeit",
           "housing":"🏠 Wohnen","clothes":"👕 Kleidung","health":"💊 Gesundheit",
           "education":"📚 Bildung","pets":"🐾 Haustiere","other":"🔧 Sonstiges"},
}
MONTHS = {
    "ru": ["","янв","фев","мар","апр","май","июн","июл","авг","сен","окт","ноя","дек"],
    "uk": ["","січ","лют","бер","кві","тра","чер","лип","сер","вер","жов","лис","гру"],
    "en": ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],
    "de": ["","Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"],
}

# ── Переклади ──────────────────────────────────────────────────────────────────
T = {
"uk": {
    "btn_add":"➕ Додати витрату","btn_quick":"⚡ Швидко",
    "btn_today":"📅 За сьогодні","btn_week":"📆 За тиждень","btn_month":"🗓 За місяць",
    "btn_top_cat":"🏆 Топ категорій","btn_top_items":"🏷 Топ товарів",
    "btn_limit":"⚠️ Встановити ліміт","btn_delete":"🗑 Видалити останню",
    "btn_lang":"🌐 Мова","btn_currency":"💱 Валюта",
    "btn_convert":"🔄 Конвертер","btn_export":"📊 Excel звіт",
    "btn_my_cats":"⚙️ Мої категорії","btn_cancel":"❌ Скасувати",
    "btn_tmpl_add":"➕ Новий шаблон","btn_tmpl_del":"🗑 Видалити шаблон",
    "btn_ccat_add":"➕ Нова категорія","btn_ccat_del":"🗑 Видалити категорію",
    "btn_donate":"☕ Підтримати",
    "btn_income":"📈 Дохід","btn_balance":"💰 Баланс","btn_help":"📖 Інструкція",
    "income_enter_amount":"📈 <b>Додати дохід</b>\n\nВведи суму ({sym}):",
    "income_enter_source":"З якого джерела? (Зарплата, Фріланс, Подарунок...):",
    "income_saved":"✅ Дохід записано!\n  <b>{source}</b> — {amount:.2f} {sym}",
    "balance_title":"💰 <b>Баланс за {month}</b>",
    "balance_income":"📈 Доходи: <b>{amount:.2f} {sym}</b>",
    "balance_expenses":"📉 Витрати: <b>{amount:.2f} {sym}</b>",
    "balance_result_pos":"✅ Залишок: <b>+{amount:.2f} {sym}</b>",
    "balance_result_neg":"⚠️ Перевитрата: <b>-{amount:.2f} {sym}</b>",
    "balance_no_data":"Немає даних за поточний місяць.",
    "help_text":"📖 <b>Як користуватись ботом</b>\n\n"
        "➕ <b>Додати витрату</b>\nНатисни кнопку → обери категорію → введи суму → назву.\n\n"
        "⚡ <b>Швидко (Шаблони)</b>\nЗбережи часті витрати і додавай в один клік.\n\n"
        "📈 <b>Дохід</b>\nЗапиши зарплату або інші надходження. Бот покаже баланс.\n\n"
        "💰 <b>Баланс</b>\nДоходи мінус витрати за поточний місяць.\n\n"
        "📅📆🗓 <b>Звіти</b>\nВитрати за день, тиждень або місяць з розбивкою по днях.\n\n"
        "⚠️ <b>Ліміти</b>\nВстанови максимум для категорії — бот попередить при перевищенні.\n\n"
        "🔁 <b>Регулярні витрати</b>\nОренда, підписки — бот запише автоматично потрібного числа.\n\n"
        "🔔 <b>Нагадування</b>\nНе забудь занести витрати: щоденні або тижневі нагадування.\n\n"
        "📊 <b>Порівняння</b>\nПорівняй цей місяць з попереднім по категоріях.\n\n"
        "🔄 <b>Конвертер</b>\nКонвертуй будь-які валюти: <code>100 USD UAH</code>\n\n"
        "📊 <b>Excel звіт</b>\nВивантаж всі дані у таблицю.\n\n"
        "⚙️ <b>Категорії</b>\nДодай свої власні категорії.\n\n"
        "💬 <b>Зворотній зв\'язок</b>\nПишіть пропозиції або питання — відповімо!",
    "btn_finance":"💰 Фінанси","btn_reports":"📊 Звіти","btn_more":"⚙️ Більше",
    "btn_back":"⬅️ Назад",
    "btn_add_income_short":"➕ Додати дохід",
    "finance_title":"💰 <b>Фінанси за {month}</b>",
    "finance_no_income":"📈 Доходів не записано",
    "finance_add_income_hint":"Натисни «➕ Додати дохід» щоб записати надходження.",
    "btn_settings":"⚙️ Налаштування",
    "btn_feedback":"💬 Зворотній зв'язок",
    "feedback_cooldown":"⏳ Ти вже надсилав повідомлення нещодавно. Спробуй через годину.",
    "feedback_reply":"📩 <b>Відповідь від адміністратора:</b>\n\n{text}",
    "feedback_reply_sent":"✅ Відповідь надіслана користувачу {uid}.",
    "feedback_reply_fail":"❌ Не вдалось надіслати. Можливо, користувач заблокував бота.",
    "feedback_reply_usage":"Використання: /reply USER_ID текст повідомлення",

    "feedback_prompt":"💬 Напиши своє повідомлення — скаргу, пропозицію або питання.\n\nВоно буде відправлено адміністратору.",
    "feedback_sent":"✅ Повідомлення відправлено! Дякуємо за відгук.",
    "feedback_received":"📨 <b>Нове повідомлення</b>\nВід: {name} (ID: <code>{user_id}</code>)\n\n{text}",
    "settings_title":"⚙️ <b>Налаштування</b>\n\nОбери що змінити:",
    "onboard_welcome":"👋 <b>Привіт! Це My Expenses Journal.</b>\n\nСпочатку налаштуємо мову та валюту.",
    "onboard_lang":"🌐 Обери мову:",
    "onboard_cur_pri":"💱 Обери основну валюту (в ній ти будеш вводити витрати):",
    "onboard_cur_sec":"💱 Обери валюту для конвертації (буде показуватись поруч із сумою):",
    "onboard_done":"✅ Готово! Ось що вміє цей бот:\n\n""➕ <b>Додати витрату</b> — категорія, сума, назва\n""⚡ <b>Швидко</b> — шаблони для частих витрат\n""📅📆🗓 <b>Звіти</b> — за день, тиждень, місяць\n""🏆🏷 <b>Топ</b> категорій та товарів\n""⚠️ <b>Ліміти</b> — попередження при перевищенні\n""🔄 <b>Конвертер</b> валют\n""📊 <b>Excel звіт</b> для аналізу\n""🔁 <b>Регулярні</b> — авто-запис щомісяця\n""🔔 <b>Нагадування</b> — щоденні та тижневі\n""📊 <b>Порівняння</b> місяців\n""⚙️ <b>Категорії</b> — власні категорії\n\n""Починай! 👇",
    "welcome":"👋 <b>My Expenses Journal</b>\n\nОбери дію:",
    "choose_menu":"Використовуй кнопки меню 👇","cancelled":"Скасовано.",
    "choose_cat":"Обери категорію:",
    "enter_amount":"Категорія: <b>{cat}</b>\n\nВведи суму ({sym}):",
    "bad_amount":"Введи коректну суму, наприклад: 125.50\nМаксимум: 999 999",
    "enter_name":"Введи назву товару/послуги:",
    "saved":"✅ Записано!\n  <b>{name}</b> — {amount:.2f} {sym}\n  Категорія: {cat}",
    "no_data":"Немає даних за цей період.","nothing_del":"Нічого видаляти.",
    "deleted":"🗑 Видалено: <b>{name}</b> — {amount:.2f} {sym} [{cat}]",
    "limit_choose":"Обери категорію для ліміту (на місяць):",
    "limit_enter":"Ліміт для «{cat}».\nВведи суму ({sym}) (0 — зняти):",
    "limit_set":"✅ Ліміт для «{cat}»: <b>{amount:.2f} {sym}/міс</b>",
    "limit_removed":"🗑 Ліміт для «{cat}» знято.",
    "limit_over":"⚠️ <b>Ліміт по «{cat}» перевищено!</b>\nВитрачено: <b>{spent:.2f} {sym}</b> / ліміт {limit:.2f} {sym}",
    "limit_warn":"⚡ Використано <b>{pct:.0f}%</b> ліміту по «{cat}» ({spent:.2f} / {limit:.2f} {sym})",
    "lang_switched":"✅ Мову змінено на <b>Українська</b> 🇺🇦",
    "total_label":"Всього","times":"раз",
    "title_today":"📅 Сьогодні","title_week":"📆 Цей тиждень","title_month":"🗓 Цей місяць",
    "title_top_cat":"🏆 <b>Топ категорій (місяць)</b>",
    "title_top_items":"🏷 <b>Топ товарів/послуг (місяць)</b>",
    "tmpl_none":"Шаблонів ще немає.\nНатисни «➕ Новий шаблон».",
    "tmpl_choose":"Обери шаблон або дію:",
    "tmpl_add_name":"Введи назву шаблону (наприклад: Трамвай, Кава):",
    "tmpl_add_amount":"Сума для «{name}» ({sym}):",
    "tmpl_add_cat":"Категорія для «{name}»:",
    "tmpl_added":"✅ Шаблон «{name}» збережено ({amount:.2f} {sym}, {cat})",
    "tmpl_del_choose":"Який шаблон видалити?","tmpl_deleted":"🗑 Шаблон «{name}» видалено.",
    "tmpl_no_del":"Немає шаблонів для видалення.",
    "lang_select":"🌐 Обери мову:",
    "curr_primary":"Обери <b>основну валюту</b>\n(в ній вводяться всі витрати):",
    "curr_secondary":"Обери <b>валюту конвертації</b>\n(відображається поруч із сумою):",
    "curr_none":"🚫 Без конвертації",
    "curr_set_primary":"✅ Основна валюта: <b>{cur}</b> {sym}",
    "curr_set_secondary":"✅ Конвертація: <b>{cur}</b> {sym}",
    "curr_set_secondary_none":"✅ Конвертацію вимкнено.",
    "convert_prompt":"Введи суму та валюти в одному повідомленні:\n\n<code>12000 UAH USD</code>\n<code>100 EUR PLN</code>\n<code>50 GBP UAH</code>\n\nПідтримуються всі світові коди (USD, EUR, GBP, PLN, CZK, ...)",
    "convert_result":"💱 <b>{amount:.2f} {from_cur}</b> = <b>{result:.2f} {to_cur}</b>\n📈 Курс: 1 {from_cur} = {rate:.4f} {to_cur}",
    "convert_error":"❌ Не вдалось конвертувати.\n\nФормат: <code>12000 UAH USD</code>\nПеревір коди валют (3 латинські літери).",
    "convert_unavail":"❌ Курс для <b>{cur}</b> недоступний. Перевір код валюти.",
    "export_period":"Обери період для Excel звіту:",
    "export_btn_today":"📅 Сьогодні","export_btn_week":"📆 Цей тиждень",
    "export_btn_month":"🗓 Цей місяць","export_btn_all":"📁 Всі записи",
    "export_empty":"Немає даних для звіту.",
    "export_sending":"⏳ Формую файл...",
    "export_no_xlsx":"Встанови openpyxl:\n<code>pip install openpyxl</code>",
    "export_sheet_all":"Витрати","export_sheet_cat":"По категоріях","export_sheet_day":"По днях",
    "ccat_header":"⚙️ <b>Мої категорії</b>",
    "ccat_list_item":"• {label}",
    "ccat_none":"Власних категорій ще немає.",
    "ccat_action":"Що зробити?",
    "ccat_add_name":"Введи назву нової категорії.\nМожна з емодзі: <code>🚕 Таксі</code>",
    "ccat_too_long":"Назва занадто довга (макс. 40 символів).",
    "ccat_added":"✅ Категорію «{name}» додано!",
    "ccat_del_choose":"Яку категорію видалити?",
    "ccat_deleted":"🗑 Категорію «{name}» видалено.",
    "ccat_no_del":"Немає власних категорій для видалення.",
    "btn_recurring":"🔁 Регулярні","btn_reminders":"🔔 Нагадування","btn_compare":"📊 Порівняння",
    "recur_none":"Регулярних витрат немає.","recur_choose":"Оберіть або додайте:",
    "btn_recur_add":"➕ Додати","btn_recur_del":"🗑 Видалити",
    "recur_add_name":"Назва (наприклад: Оренда, Netflix):","recur_add_amount":"Сума для «{name}» ({sym}):",
    "recur_add_cat":"Категорія для «{name}»:","recur_add_day":"День місяця (1–31)\n(якщо місяць коротший — спрацює в останній день):",
    "recur_bad_day":"Введи число від 1 до 31.",
    "recur_added":"✅ «{name}» додано!\nЩомісяця {day}-го: {amount:.2f} {sym}",
    "recur_del_choose":"Який запис видалити?","recur_deleted":"🗑 «{name}» видалено.","recur_no_del":"Нічого видаляти.",
    "recur_auto":"🔁 Авто-запис: <b>{name}</b> — {amount:.2f} {sym}",
    "remind_title":"🔔 <b>Нагадування</b>",
    "remind_current":"Поточні:\n• Неактивність: {inact} дн.\n• Щоденний: {daily}\n• Тижневий: {weekly} (день {wday})",
    "remind_off":"вимк.","remind_btn_inactive":"⏰ Неактивність","remind_btn_daily":"📅 Щоденний","remind_btn_weekly":"📆 Тижневий",
    "remind_inactive_prompt":"Через скільки днів без витрат нагадати?\n(0 = вимкнути, макс. 30):",
    "remind_daily_prompt":"Час щоденного нагадування (HH:MM UTC) або 0:","remind_weekly_day_prompt":"День тижня (0=Пн..6=Нд) або -1:",
    "remind_weekly_time_prompt":"Час тижневого нагадування (HH:MM UTC):",
    "remind_saved":"✅ Нагадування збережено.","remind_bad_time":"Формат: HH:MM (наприклад: 09:00)",
    "remind_bad_inactive":"Число від 0 до 30.","remind_bad_wday":"Число від -1 до 6.",
    "remind_inact_msg":"👋 Ти не записував витрати вже <b>{days} дн.</b> Не забудь!",
    "remind_daily_msg":"📅 Час перевірити витрати! Натисни «{btn}».",
    "remind_weekly_msg":"📆 Тиждень закінчується! Натисни «{btn}».",
    "compare_title":"📊 <b>Порівняння місяців</b>","compare_this":"Цей місяць","compare_prev":"Минулий місяць",
    "compare_diff_more":"▲ +{diff:.2f} {sym} (+{pct:.0f}%) більше","compare_diff_less":"▼ {diff:.2f} {sym} ({pct:.0f}%) менше",
    "compare_diff_same":"◆ Без змін","compare_no_data":"Недостатньо даних.","compare_by_cat":"По категоріях:",
    "donate_msg":"☕ <b>Підтримати проєкт</b>\n\nЯкщо бот корисний — буду вдячний за каву!\n\n{url}",
    "donate_no_url":"Посилання для донату не налаштоване.",
},
"ru": {
    "btn_add":"➕ Добавить трату","btn_quick":"⚡ Быстро",
    "btn_today":"📅 За сегодня","btn_week":"📆 За неделю","btn_month":"🗓 За месяц",
    "btn_top_cat":"🏆 Топ категорий","btn_top_items":"🏷 Топ товаров",
    "btn_limit":"⚠️ Установить лимит","btn_delete":"🗑 Удалить последнюю",
    "btn_lang":"🌐 Язык","btn_currency":"💱 Валюта",
    "btn_convert":"🔄 Конвертер","btn_export":"📊 Excel отчёт",
    "btn_my_cats":"⚙️ Мои категории","btn_cancel":"❌ Отмена",
    "btn_tmpl_add":"➕ Новый шаблон","btn_tmpl_del":"🗑 Удалить шаблон",
    "btn_ccat_add":"➕ Новая категория","btn_ccat_del":"🗑 Удалить категорию",
    "btn_donate":"☕ Поддержать",
    "btn_income":"📈 Доход","btn_balance":"💰 Баланс","btn_help":"📖 Инструкция",
    "income_enter_amount":"📈 <b>Добавить доход</b>\n\nВведи сумму ({sym}):",
    "income_enter_source":"Источник дохода? (Зарплата, Фриланс, Подарок...):",
    "income_saved":"✅ Доход записан!\n  <b>{source}</b> — {amount:.2f} {sym}",
    "balance_title":"💰 <b>Баланс за {month}</b>",
    "balance_income":"📈 Доходы: <b>{amount:.2f} {sym}</b>",
    "balance_expenses":"📉 Расходы: <b>{amount:.2f} {sym}</b>",
    "balance_result_pos":"✅ Остаток: <b>+{amount:.2f} {sym}</b>",
    "balance_result_neg":"⚠️ Перерасход: <b>-{amount:.2f} {sym}</b>",
    "balance_no_data":"Нет данных за текущий месяц.",
    "help_text":"📖 <b>Как пользоваться ботом</b>\n\n"
        "➕ <b>Добавить трату</b>\nНажми кнопку → выбери категорию → введи сумму → название.\n\n"
        "⚡ <b>Быстро (Шаблоны)</b>\nСохрани частые траты и добавляй в один клик.\n\n"
        "📈 <b>Доход</b>\nЗапиши зарплату или другие поступления. Бот покажет баланс.\n\n"
        "💰 <b>Баланс</b>\nДоходы минус расходы за текущий месяц.\n\n"
        "📅📆🗓 <b>Отчёты</b>\nРасходы за день, неделю или месяц с разбивкой по дням.\n\n"
        "⚠️ <b>Лимиты</b>\nУстанови максимум для категории — бот предупредит при превышении.\n\n"
        "🔁 <b>Регулярные расходы</b>\nАренда, подписки — бот запишет автоматически нужного числа.\n\n"
        "🔔 <b>Напоминания</b>\nЕжедневные или еженедельные напоминания записать траты.\n\n"
        "📊 <b>Сравнение</b>\nСравни этот месяц с прошлым по категориям.\n\n"
        "🔄 <b>Конвертер</b>\nКонвертируй любые валюты: <code>100 USD UAH</code>\n\n"
        "📊 <b>Excel отчёт</b>\nВыгрузи все данные в таблицу.\n\n"
        "⚙️ <b>Категории</b>\nДобавь свои собственные категории.\n\n"
        "💬 <b>Обратная связь</b>\nПишите предложения или вопросы — ответим!",
    "btn_finance":"💰 Финансы","btn_reports":"📊 Отчёты","btn_more":"⚙️ Ещё",
    "btn_back":"⬅️ Назад",
    "btn_add_income_short":"➕ Добавить доход",
    "finance_title":"💰 <b>Финансы за {month}</b>",
    "finance_no_income":"📈 Доходов не записано",
    "finance_add_income_hint":"Нажми «➕ Добавить доход» чтобы записать поступления.",
    "btn_settings":"⚙️ Настройки",
    "btn_feedback":"💬 Обратная связь",
    "feedback_cooldown":"⏳ Ты уже отправлял сообщение недавно. Попробуй через час.",
    "feedback_reply":"📩 <b>Ответ от администратора:</b>\n\n{text}",
    "feedback_reply_sent":"✅ Ответ отправлен пользователю {uid}.",
    "feedback_reply_fail":"❌ Не удалось отправить. Возможно, пользователь заблокировал бота.",
    "feedback_reply_usage":"Использование: /reply USER_ID текст сообщения",

    "feedback_prompt":"💬 Напиши своё сообщение — жалобу, предложение или вопрос.\n\nОно будет отправлено администратору.",
    "feedback_sent":"✅ Сообщение отправлено! Спасибо за отзыв.",
    "feedback_received":"📨 <b>Новое сообщение</b>\nОт: {name} (ID: <code>{user_id}</code>)\n\n{text}",
    "settings_title":"⚙️ <b>Настройки</b>\n\nВыбери что изменить:",
    "onboard_welcome":"👋 <b>Привет! Это My Expenses Journal.</b>\n\nСначала настроим язык и валюту.",
    "onboard_lang":"🌐 Выбери язык:",
    "onboard_cur_pri":"💱 Выбери основную валюту (в ней будешь вводить расходы):",
    "onboard_cur_sec":"💱 Выбери валюту для конвертации (будет показываться рядом с суммой):",
    "onboard_done":"✅ Готово! Вот что умеет этот бот:\n\n""➕ <b>Добавить трату</b> — категория, сумма, название\n""⚡ <b>Быстро</b> — шаблоны для частых расходов\n""📅📆🗓 <b>Отчёты</b> — за день, неделю, месяц\n""🏆🏷 <b>Топ</b> категорий и товаров\n""⚠️ <b>Лимиты</b> — предупреждения при превышении\n""🔄 <b>Конвертер</b> валют\n""📊 <b>Excel отчёт</b> для анализа\n""🔁 <b>Регулярные</b> — авто-запись каждый месяц\n""🔔 <b>Напоминания</b> — ежедневные и еженедельные\n""📊 <b>Сравнение</b> месяцев\n""⚙️ <b>Категории</b> — свои категории\n\n""Начинай! 👇",
    "welcome":"👋 <b>My Expenses Journal</b>\n\nВыбери действие:",
    "choose_menu":"Используй кнопки меню 👇","cancelled":"Отменено.",
    "choose_cat":"Выбери категорию:",
    "enter_amount":"Категория: <b>{cat}</b>\n\nВведи сумму ({sym}):",
    "bad_amount":"Введи корректную сумму, например: 125.50\nМаксимум: 999 999",
    "enter_name":"Введи название товара/услуги:",
    "saved":"✅ Записано!\n  <b>{name}</b> — {amount:.2f} {sym}\n  Категория: {cat}",
    "no_data":"Нет данных за этот период.","nothing_del":"Нечего удалять.",
    "deleted":"🗑 Удалено: <b>{name}</b> — {amount:.2f} {sym} [{cat}]",
    "limit_choose":"Выбери категорию для лимита (на месяц):",
    "limit_enter":"Лимит для «{cat}».\nВведи сумму ({sym}) (0 — снять):",
    "limit_set":"✅ Лимит для «{cat}»: <b>{amount:.2f} {sym}/мес</b>",
    "limit_removed":"🗑 Лимит для «{cat}» снят.",
    "limit_over":"⚠️ <b>Лимит по «{cat}» превышен!</b>\nПотрачено: <b>{spent:.2f} {sym}</b> / лимит {limit:.2f} {sym}",
    "limit_warn":"⚡ Использовано <b>{pct:.0f}%</b> лимита по «{cat}» ({spent:.2f} / {limit:.2f} {sym})",
    "lang_switched":"✅ Язык изменён на <b>Русский</b> 🇷🇺",
    "total_label":"Всего","times":"раз",
    "title_today":"📅 Сегодня","title_week":"📆 Эта неделя","title_month":"🗓 Этот месяц",
    "title_top_cat":"🏆 <b>Топ категорий (месяц)</b>",
    "title_top_items":"🏷 <b>Топ товаров/услуг (месяц)</b>",
    "tmpl_none":"Шаблонов ещё нет.\nНажми «➕ Новый шаблон».",
    "tmpl_choose":"Выбери шаблон или действие:",
    "tmpl_add_name":"Введи название шаблона (например: Трамвай, Кофе):",
    "tmpl_add_amount":"Сумма для «{name}» ({sym}):",
    "tmpl_add_cat":"Категория для «{name}»:",
    "tmpl_added":"✅ Шаблон «{name}» сохранён ({amount:.2f} {sym}, {cat})",
    "tmpl_del_choose":"Какой шаблон удалить?","tmpl_deleted":"🗑 Шаблон «{name}» удалён.",
    "tmpl_no_del":"Нет шаблонов для удаления.",
    "lang_select":"🌐 Выбери язык:",
    "curr_primary":"Выбери <b>основную валюту</b>\n(в ней вводятся все траты):",
    "curr_secondary":"Выбери <b>валюту конвертации</b>\n(показывается рядом с суммой):",
    "curr_none":"🚫 Без конвертации",
    "curr_set_primary":"✅ Основная валюта: <b>{cur}</b> {sym}",
    "curr_set_secondary":"✅ Конвертация: <b>{cur}</b> {sym}",
    "curr_set_secondary_none":"✅ Конвертация отключена.",
    "convert_prompt":"Введи сумму и валюты в одном сообщении:\n\n<code>12000 UAH USD</code>\n<code>100 EUR PLN</code>\n<code>50 GBP UAH</code>\n\nПоддерживаются все мировые коды (USD, EUR, GBP, PLN, CZK, ...)",
    "convert_result":"💱 <b>{amount:.2f} {from_cur}</b> = <b>{result:.2f} {to_cur}</b>\n📈 Курс: 1 {from_cur} = {rate:.4f} {to_cur}",
    "convert_error":"❌ Не удалось конвертировать.\n\nФормат: <code>12000 UAH USD</code>\nПроверь коды валют (3 латинские буквы).",
    "convert_unavail":"❌ Курс для <b>{cur}</b> недоступен. Проверь код валюты.",
    "export_period":"Выбери период для Excel отчёта:",
    "export_btn_today":"📅 Сегодня","export_btn_week":"📆 Эта неделя",
    "export_btn_month":"🗓 Этот месяц","export_btn_all":"📁 Все записи",
    "export_empty":"Нет данных для отчёта.",
    "export_sending":"⏳ Формирую файл...",
    "export_no_xlsx":"Установи openpyxl:\n<code>pip install openpyxl</code>",
    "export_sheet_all":"Расходы","export_sheet_cat":"По категориям","export_sheet_day":"По дням",
    "ccat_header":"⚙️ <b>Мои категории</b>",
    "ccat_list_item":"• {label}",
    "ccat_none":"Своих категорий ещё нет.",
    "ccat_action":"Что сделать?",
    "ccat_add_name":"Введи название новой категории.\nМожно с эмодзи: <code>🚕 Такси</code>",
    "ccat_too_long":"Название слишком длинное (макс. 40 символов).",
    "ccat_added":"✅ Категория «{name}» добавлена!",
    "ccat_del_choose":"Какую категорию удалить?",
    "ccat_deleted":"🗑 Категория «{name}» удалена.",
    "ccat_no_del":"Нет своих категорий для удаления.",
    "btn_recurring":"🔁 Регулярные","btn_reminders":"🔔 Напоминания","btn_compare":"📊 Сравнение",
    "recur_none":"Регулярных расходов нет.","recur_choose":"Выбери или добавь:",
    "btn_recur_add":"➕ Добавить","btn_recur_del":"🗑 Удалить",
    "recur_add_name":"Название (Аренда, Netflix):","recur_add_amount":"Сумма для «{name}» ({sym}):",
    "recur_add_cat":"Категория для «{name}»:","recur_add_day":"День месяца (1–31)\n(если месяц короче — сработает в последний день):",
    "recur_bad_day":"Введи число от 1 до 31.",
    "recur_added":"✅ «{name}» добавлен!\nКаждый месяц {day}-го: {amount:.2f} {sym}",
    "recur_del_choose":"Какой расход удалить?","recur_deleted":"🗑 «{name}» удалён.","recur_no_del":"Нечего удалять.",
    "recur_auto":"🔁 Авто-запись: <b>{name}</b> — {amount:.2f} {sym}",
    "remind_title":"🔔 <b>Напоминания</b>",
    "remind_current":"Текущие:\n• Неактивность: {inact} дн.\n• Ежедневный: {daily}\n• Еженедельный: {weekly} (день {wday})",
    "remind_off":"выкл.","remind_btn_inactive":"⏰ Неактивность","remind_btn_daily":"📅 Ежедневный","remind_btn_weekly":"📆 Еженедельный",
    "remind_inactive_prompt":"Через сколько дней без расходов напомнить?\n(0 = выключить, макс. 30):",
    "remind_daily_prompt":"Время ежедневного напоминания (HH:MM UTC) или 0:","remind_weekly_day_prompt":"День недели (0=Пн..6=Вс) или -1:",
    "remind_weekly_time_prompt":"Время еженедельного напоминания (HH:MM UTC):",
    "remind_saved":"✅ Напоминание сохранено.","remind_bad_time":"Формат: HH:MM (например: 09:00)",
    "remind_bad_inactive":"Число от 0 до 30.","remind_bad_wday":"Число от -1 до 6.",
    "remind_inact_msg":"👋 Ты не записывал расходы уже <b>{days} дн.</b> Не забудь!",
    "remind_daily_msg":"📅 Время проверить расходы! Нажми «{btn}».",
    "remind_weekly_msg":"📆 Неделя заканчивается! Нажми «{btn}».",
    "compare_title":"📊 <b>Сравнение месяцев</b>","compare_this":"Этот месяц","compare_prev":"Прошлый месяц",
    "compare_diff_more":"▲ +{diff:.2f} {sym} (+{pct:.0f}%) больше","compare_diff_less":"▼ {diff:.2f} {sym} ({pct:.0f}%) меньше",
    "compare_diff_same":"◆ Без изменений","compare_no_data":"Недостаточно данных.","compare_by_cat":"По категориям:",
    "donate_msg":"☕ <b>Поддержать проект</b>\n\nЕсли бот полезен — буду благодарен за кофе!\n\n{url}",
    "donate_no_url":"Ссылка для доната не настроена.",
},
"en": {
    "btn_add":"➕ Add expense","btn_quick":"⚡ Quick",
    "btn_today":"📅 Today","btn_week":"📆 This week","btn_month":"🗓 This month",
    "btn_top_cat":"🏆 Top categories","btn_top_items":"🏷 Top items",
    "btn_limit":"⚠️ Set limit","btn_delete":"🗑 Delete last",
    "btn_lang":"🌐 Language","btn_currency":"💱 Currency",
    "btn_convert":"🔄 Converter","btn_export":"📊 Excel report",
    "btn_my_cats":"⚙️ My categories","btn_cancel":"❌ Cancel",
    "btn_tmpl_add":"➕ New template","btn_tmpl_del":"🗑 Delete template",
    "btn_ccat_add":"➕ New category","btn_ccat_del":"🗑 Delete category",
    "btn_donate":"☕ Donate",
    "btn_income":"📈 Income","btn_balance":"💰 Balance","btn_help":"📖 Help",
    "income_enter_amount":"📈 <b>Add income</b>\n\nEnter amount ({sym}):",
    "income_enter_source":"Income source? (Salary, Freelance, Gift...):",
    "income_saved":"✅ Income recorded!\n  <b>{source}</b> — {amount:.2f} {sym}",
    "balance_title":"💰 <b>Balance for {month}</b>",
    "balance_income":"📈 Income: <b>{amount:.2f} {sym}</b>",
    "balance_expenses":"📉 Expenses: <b>{amount:.2f} {sym}</b>",
    "balance_result_pos":"✅ Remaining: <b>+{amount:.2f} {sym}</b>",
    "balance_result_neg":"⚠️ Overspent: <b>-{amount:.2f} {sym}</b>",
    "balance_no_data":"No data for the current month.",
    "help_text":"📖 <b>How to use this bot</b>\n\n"
        "➕ <b>Add expense</b>\nTap button → choose category → enter amount → name.\n\n"
        "⚡ <b>Quick (Templates)</b>\nSave frequent expenses and add in one tap.\n\n"
        "📈 <b>Income</b>\nRecord your salary or other income. Bot shows your balance.\n\n"
        "💰 <b>Balance</b>\nIncome minus expenses for the current month.\n\n"
        "📅📆🗓 <b>Reports</b>\nExpenses for today, week or month with daily breakdown.\n\n"
        "⚠️ <b>Limits</b>\nSet a max for a category — bot warns when exceeded.\n\n"
        "🔁 <b>Recurring</b>\nRent, subscriptions — bot records automatically on the set date.\n\n"
        "🔔 <b>Reminders</b>\nDaily or weekly reminders to log expenses.\n\n"
        "📊 <b>Compare</b>\nCompare this month vs last month by category.\n\n"
        "🔄 <b>Converter</b>\nConvert any currencies: <code>100 USD UAH</code>\n\n"
        "📊 <b>Excel report</b>\nExport all data to a spreadsheet.\n\n"
        "⚙️ <b>Categories</b>\nAdd your own custom categories.\n\n"
        "💬 <b>Feedback</b>\nSend suggestions or questions — we\'ll reply!",
    "btn_finance":"💰 Finances","btn_reports":"📊 Reports","btn_more":"⚙️ More",
    "btn_back":"⬅️ Back",
    "btn_add_income_short":"➕ Add income",
    "finance_title":"💰 <b>Finances for {month}</b>",
    "finance_no_income":"📈 No income recorded",
    "finance_add_income_hint":"Press «➕ Add income» to record your earnings.",
    "btn_settings":"⚙️ Settings",
    "btn_feedback":"💬 Feedback",
    "feedback_cooldown":"⏳ You already sent a message recently. Try again in an hour.",
    "feedback_reply":"📩 <b>Reply from admin:</b>\n\n{text}",
    "feedback_reply_sent":"✅ Reply sent to user {uid}.",
    "feedback_reply_fail":"❌ Could not send. The user may have blocked the bot.",
    "feedback_reply_usage":"Usage: /reply USER_ID message text",
    "feedback_prompt":"💬 Send your message — complaint, suggestion or question.\n\nIt will be forwarded to the admin.",
    "feedback_sent":"✅ Message sent! Thank you for your feedback.",
    "feedback_received":"📨 <b>New message</b>\nFrom: {name} (ID: <code>{user_id}</code>)\n\n{text}",
    "settings_title":"⚙️ <b>Settings</b>\n\nChoose what to change:",
    "onboard_welcome":"👋 <b>Hello! This is My Expenses Journal.</b>\n\nLet's set up your language and currency first.",
    "onboard_lang":"🌐 Choose your language:",
    "onboard_cur_pri":"💱 Choose your primary currency (expenses will be entered in it):",
    "onboard_cur_sec":"💱 Choose conversion currency (shown alongside amounts):",
    "onboard_done":"✅ All set! Here's what this bot can do:\n\n""➕ <b>Add expense</b> — category, amount, name\n""⚡ <b>Quick</b> — templates for frequent expenses\n""📅📆🗓 <b>Reports</b> — daily, weekly, monthly\n""🏆🏷 <b>Top</b> categories and items\n""⚠️ <b>Limits</b> — warnings when exceeded\n""🔄 <b>Converter</b> — currency converter\n""📊 <b>Excel report</b> for analysis\n""🔁 <b>Recurring</b> — auto-record monthly\n""🔔 <b>Reminders</b> — daily and weekly\n""📊 <b>Compare</b> months\n""⚙️ <b>Categories</b> — custom categories\n\n""Let's go! 👇",
    "welcome":"👋 <b>My Expenses Journal</b>\n\nChoose an action:",
    "choose_menu":"Use the menu buttons 👇","cancelled":"Cancelled.",
    "choose_cat":"Choose a category:",
    "enter_amount":"Category: <b>{cat}</b>\n\nEnter amount ({sym}):",
    "bad_amount":"Enter a valid amount, e.g. 125.50\nMaximum: 999 999",
    "enter_name":"Enter item / service name:",
    "saved":"✅ Saved!\n  <b>{name}</b> — {amount:.2f} {sym}\n  Category: {cat}",
    "no_data":"No data for this period.","nothing_del":"Nothing to delete.",
    "deleted":"🗑 Deleted: <b>{name}</b> — {amount:.2f} {sym} [{cat}]",
    "limit_choose":"Choose a category for the monthly limit:",
    "limit_enter":"Limit for «{cat}».\nEnter amount ({sym}) (0 to remove):",
    "limit_set":"✅ Limit for «{cat}»: <b>{amount:.2f} {sym}/month</b>",
    "limit_removed":"🗑 Limit for «{cat}» removed.",
    "limit_over":"⚠️ <b>Limit for «{cat}» exceeded!</b>\nSpent: <b>{spent:.2f} {sym}</b> / limit {limit:.2f} {sym}",
    "limit_warn":"⚡ <b>{pct:.0f}%</b> of limit used for «{cat}» ({spent:.2f} / {limit:.2f} {sym})",
    "lang_switched":"✅ Language changed to <b>English</b> 🇬🇧",
    "total_label":"Total","times":"×",
    "title_today":"📅 Today","title_week":"📆 This week","title_month":"🗓 This month",
    "title_top_cat":"🏆 <b>Top categories (month)</b>",
    "title_top_items":"🏷 <b>Top items (month)</b>",
    "tmpl_none":"No templates yet.\nPress «➕ New template» to create one.",
    "tmpl_choose":"Choose a template or action:",
    "tmpl_add_name":"Enter template name (e.g. Coffee, Bus):",
    "tmpl_add_amount":"Amount for «{name}» ({sym}):",
    "tmpl_add_cat":"Category for «{name}»:",
    "tmpl_added":"✅ Template «{name}» saved ({amount:.2f} {sym}, {cat})",
    "tmpl_del_choose":"Which template to delete?","tmpl_deleted":"🗑 Template «{name}» deleted.",
    "tmpl_no_del":"No templates to delete.",
    "lang_select":"🌐 Choose language:",
    "curr_primary":"Choose your <b>primary currency</b>\n(all expenses are entered in this):",
    "curr_secondary":"Choose <b>conversion currency</b>\n(shown alongside amounts):",
    "curr_none":"🚫 No conversion",
    "curr_set_primary":"✅ Primary currency: <b>{cur}</b> {sym}",
    "curr_set_secondary":"✅ Conversion: <b>{cur}</b> {sym}",
    "curr_set_secondary_none":"✅ Conversion disabled.",
    "convert_prompt":"Send amount and currencies in one message:\n\n<code>12000 UAH USD</code>\n<code>100 EUR PLN</code>\n<code>50 GBP UAH</code>\n\nAll world currency codes supported (USD, EUR, GBP, PLN, CZK, ...)",
    "convert_result":"💱 <b>{amount:.2f} {from_cur}</b> = <b>{result:.2f} {to_cur}</b>\n📈 Rate: 1 {from_cur} = {rate:.4f} {to_cur}",
    "convert_error":"❌ Could not convert.\n\nFormat: <code>12000 UAH USD</code>\nCheck currency codes (3 Latin letters).",
    "convert_unavail":"❌ Rate for <b>{cur}</b> not available. Check the currency code.",
    "export_period":"Choose period for Excel report:",
    "export_btn_today":"📅 Today","export_btn_week":"📆 This week",
    "export_btn_month":"🗓 This month","export_btn_all":"📁 All records",
    "export_empty":"No data for report.",
    "export_sending":"⏳ Generating file...",
    "export_no_xlsx":"Install openpyxl:\n<code>pip install openpyxl</code>",
    "export_sheet_all":"Expenses","export_sheet_cat":"By Category","export_sheet_day":"By Day",
    "ccat_header":"⚙️ <b>My categories</b>",
    "ccat_list_item":"• {label}",
    "ccat_none":"No custom categories yet.",
    "ccat_action":"What would you like to do?",
    "ccat_add_name":"Enter new category name.\nEmoji optional: <code>🚕 Taxi</code>",
    "ccat_too_long":"Name too long (max 40 characters).",
    "ccat_added":"✅ Category «{name}» added!",
    "ccat_del_choose":"Which category to delete?",
    "ccat_deleted":"🗑 Category «{name}» deleted.",
    "ccat_no_del":"No custom categories to delete.",
    "btn_recurring":"🔁 Recurring","btn_reminders":"🔔 Reminders","btn_compare":"📊 Compare",
    "recur_none":"No recurring expenses.","recur_choose":"Choose or add:",
    "btn_recur_add":"➕ Add","btn_recur_del":"🗑 Delete",
    "recur_add_name":"Name (Rent, Netflix):","recur_add_amount":"Amount for «{name}» ({sym}):",
    "recur_add_cat":"Category for «{name}»:","recur_add_day":"Day of month (1–31)\n(fires on the last day if month is shorter):",
    "recur_bad_day":"Enter number from 1 to 28. (1–31)",
    "recur_added":"✅ «{name}» added!\nEvery month on the {day}: {amount:.2f} {sym}",
    "recur_del_choose":"Which to delete?","recur_deleted":"🗑 «{name}» deleted.","recur_no_del":"Nothing to delete.",
    "recur_auto":"🔁 Auto-recorded: <b>{name}</b> — {amount:.2f} {sym}",
    "remind_title":"🔔 <b>Reminders</b>",
    "remind_current":"Current:\n• Inactivity: {inact} days\n• Daily: {daily}\n• Weekly: {weekly} (day {wday})",
    "remind_off":"off","remind_btn_inactive":"⏰ Inactivity","remind_btn_daily":"📅 Daily","remind_btn_weekly":"📆 Weekly",
    "remind_inactive_prompt":"After how many inactive days to remind?\n(0 = off, max 30):",
    "remind_daily_prompt":"Daily reminder time (HH:MM UTC) or 0:","remind_weekly_day_prompt":"Day of week (0=Mon..6=Sun) or -1:",
    "remind_weekly_time_prompt":"Weekly reminder time (HH:MM UTC):",
    "remind_saved":"✅ Reminder saved.","remind_bad_time":"Format: HH:MM (e.g. 09:00)",
    "remind_bad_inactive":"Number from 0 to 30.","remind_bad_wday":"Number from -1 to 6.",
    "remind_inact_msg":"👋 You haven\'t logged expenses in <b>{days} days</b>. Don\'t forget!",
    "remind_daily_msg":"📅 Time to check expenses! Press «{btn}».",
    "remind_weekly_msg":"📆 Week is ending! Press «{btn}».",
    "compare_title":"📊 <b>Month comparison</b>","compare_this":"This month","compare_prev":"Last month",
    "compare_diff_more":"▲ +{diff:.2f} {sym} (+{pct:.0f}%) more","compare_diff_less":"▼ {diff:.2f} {sym} ({pct:.0f}%) less",
    "compare_diff_same":"◆ No change","compare_no_data":"Not enough data.","compare_by_cat":"By category:",
    "donate_msg":"☕ <b>Support the project</b>\n\nIf this bot is useful, a coffee would be appreciated!\n\n{url}",
    "donate_no_url":"Donation link is not configured.",
},
"de": {
    "btn_add":"➕ Ausgabe hinzufügen","btn_quick":"⚡ Schnell",
    "btn_today":"📅 Heute","btn_week":"📆 Diese Woche","btn_month":"🗓 Diesen Monat",
    "btn_top_cat":"🏆 Top Kategorien","btn_top_items":"🏷 Top Artikel",
    "btn_limit":"⚠️ Limit setzen","btn_delete":"🗑 Letzten löschen",
    "btn_lang":"🌐 Sprache","btn_currency":"💱 Währung",
    "btn_convert":"🔄 Konverter","btn_export":"📊 Excel-Bericht",
    "btn_my_cats":"⚙️ Meine Kategorien","btn_cancel":"❌ Abbrechen",
    "btn_tmpl_add":"➕ Neue Vorlage","btn_tmpl_del":"🗑 Vorlage löschen",
    "btn_ccat_add":"➕ Neue Kategorie","btn_ccat_del":"🗑 Kategorie löschen",
    "btn_donate":"☕ Unterstützen",
    "btn_income":"📈 Einnahmen","btn_balance":"💰 Bilanz","btn_help":"📖 Anleitung",
    "income_enter_amount":"📈 <b>Einnahmen hinzufügen</b>\n\nBetrag eingeben ({sym}):",
    "income_enter_source":"Einnahmequelle? (Gehalt, Freelance, Geschenk...):",
    "income_saved":"✅ Einnahmen gespeichert!\n  <b>{source}</b> — {amount:.2f} {sym}",
    "balance_title":"💰 <b>Bilanz für {month}</b>",
    "balance_income":"📈 Einnahmen: <b>{amount:.2f} {sym}</b>",
    "balance_expenses":"📉 Ausgaben: <b>{amount:.2f} {sym}</b>",
    "balance_result_pos":"✅ Rest: <b>+{amount:.2f} {sym}</b>",
    "balance_result_neg":"⚠️ Überzogen: <b>-{amount:.2f} {sym}</b>",
    "balance_no_data":"Keine Daten für den aktuellen Monat.",
    "help_text":"📖 <b>Bot-Anleitung</b>\n\n"
        "➕ <b>Ausgabe hinzufügen</b>\nTippe Taste → Kategorie → Betrag → Name.\n\n"
        "⚡ <b>Schnell (Vorlagen)</b>\nHäufige Ausgaben speichern und per Tipp hinzufügen.\n\n"
        "📈 <b>Einnahmen</b>\nGehalt oder andere Einnahmen erfassen. Bot zeigt Bilanz.\n\n"
        "💰 <b>Bilanz</b>\nEinnahmen minus Ausgaben für den aktuellen Monat.\n\n"
        "📅📆🗓 <b>Berichte</b>\nAusgaben für heute, Woche oder Monat mit Tagesübersicht.\n\n"
        "⚠️ <b>Limits</b>\nMaximum für eine Kategorie setzen — Bot warnt bei Überschreitung.\n\n"
        "🔁 <b>Wiederkehrend</b>\nMiete, Abos — Bot bucht automatisch am eingestellten Tag.\n\n"
        "🔔 <b>Erinnerungen</b>\nTägliche oder wöchentliche Erinnerungen zum Erfassen.\n\n"
        "📊 <b>Vergleich</b>\nDieser Monat vs. letzter Monat nach Kategorien.\n\n"
        "🔄 <b>Konverter</b>\nWährungen umrechnen: <code>100 USD UAH</code>\n\n"
        "📊 <b>Excel-Bericht</b>\nAlle Daten als Tabelle exportieren.\n\n"
        "⚙️ <b>Kategorien</b>\nEigene Kategorien hinzufügen.\n\n"
        "💬 <b>Feedback</b>\nVorschläge oder Fragen — wir antworten!",
    "btn_finance":"💰 Finanzen","btn_reports":"📊 Berichte","btn_more":"⚙️ Mehr",
    "btn_back":"⬅️ Zurück",
    "btn_add_income_short":"➕ Einnahmen hinzufügen",
    "finance_title":"💰 <b>Finanzen für {month}</b>",
    "finance_no_income":"📈 Keine Einnahmen erfasst",
    "finance_add_income_hint":"Drücke «➕ Einnahmen hinzufügen» um Einnahmen zu erfassen.",
    "btn_settings":"⚙️ Einstellungen",
    "btn_feedback":"💬 Feedback",
    "feedback_cooldown":"⏳ Du hast kürzlich eine Nachricht gesendet. Versuche es in einer Stunde.",
    "feedback_reply":"📩 <b>Antwort vom Admin:</b>\n\n{text}",
    "feedback_reply_sent":"✅ Antwort an Nutzer {uid} gesendet.",
    "feedback_reply_fail":"❌ Senden fehlgeschlagen. Der Nutzer hat den Bot möglicherweise blockiert.",
    "feedback_reply_usage":"Verwendung: /reply USER_ID Nachrichtentext",
    "feedback_prompt":"💬 Schreib deine Nachricht — Beschwerde, Vorschlag oder Frage.\n\nSie wird an den Admin weitergeleitet.",
    "feedback_sent":"✅ Nachricht gesendet! Danke für dein Feedback.",
    "feedback_received":"📨 <b>Neue Nachricht</b>\nVon: {name} (ID: <code>{user_id}</code>)\n\n{text}",
    "settings_title":"⚙️ <b>Einstellungen</b>\n\nWas möchtest du ändern?",
    "onboard_welcome":"👋 <b>Hallo! Das ist My Expenses Journal.</b>\n\nZuerst richten wir Sprache und Währung ein.",
    "onboard_lang":"🌐 Sprache wählen:",
    "onboard_cur_pri":"💱 Hauptwährung wählen (Ausgaben werden darin eingegeben):",
    "onboard_cur_sec":"💱 Umrechnungswährung wählen (neben Beträgen angezeigt):",
    "onboard_done":"✅ Fertig! Das kann dieser Bot:\n\n""➕ <b>Ausgabe hinzufügen</b> — Kategorie, Betrag, Name\n""⚡ <b>Schnell</b> — Vorlagen für häufige Ausgaben\n""📅📆🗓 <b>Berichte</b> — täglich, wöchentlich, monatlich\n""🏆🏷 <b>Top</b> Kategorien und Artikel\n""⚠️ <b>Limits</b> — Warnungen bei Überschreitung\n""🔄 <b>Konverter</b> — Währungsrechner\n""📊 <b>Excel-Bericht</b> zur Analyse\n""🔁 <b>Wiederkehrend</b> — Auto-Buchung monatlich\n""🔔 <b>Erinnerungen</b> — täglich und wöchentlich\n""📊 <b>Vergleich</b> — Monatsvergleich\n""⚙️ <b>Kategorien</b> — eigene Kategorien\n\n""Los geht's! 👇",
    "welcome":"👋 <b>My Expenses Journal</b>\n\nAktion wählen:",
    "choose_menu":"Menütasten verwenden 👇","cancelled":"Abgebrochen.",
    "choose_cat":"Kategorie wählen:",
    "enter_amount":"Kategorie: <b>{cat}</b>\n\nBetrag eingeben ({sym}):",
    "bad_amount":"Gültigen Betrag eingeben, z.B. 125.50\nMaximum: 999 999",
    "enter_name":"Artikel-/Dienstleistungsname eingeben:",
    "saved":"✅ Gespeichert!\n  <b>{name}</b> — {amount:.2f} {sym}\n  Kategorie: {cat}",
    "no_data":"Keine Daten für diesen Zeitraum.","nothing_del":"Nichts zu löschen.",
    "deleted":"🗑 Gelöscht: <b>{name}</b> — {amount:.2f} {sym} [{cat}]",
    "limit_choose":"Kategorie für das monatliche Limit wählen:",
    "limit_enter":"Limit für «{cat}».\nBetrag ({sym}) eingeben (0 = entfernen):",
    "limit_set":"✅ Limit für «{cat}»: <b>{amount:.2f} {sym}/Monat</b>",
    "limit_removed":"🗑 Limit für «{cat}» entfernt.",
    "limit_over":"⚠️ <b>Limit für «{cat}» überschritten!</b>\nAusgegeben: <b>{spent:.2f} {sym}</b> / Limit {limit:.2f} {sym}",
    "limit_warn":"⚡ <b>{pct:.0f}%</b> des Limits für «{cat}» ({spent:.2f} / {limit:.2f} {sym})",
    "lang_switched":"✅ Sprache auf <b>Deutsch</b> geändert 🇩🇪",
    "total_label":"Gesamt","times":"×",
    "title_today":"📅 Heute","title_week":"📆 Diese Woche","title_month":"🗓 Diesen Monat",
    "title_top_cat":"🏆 <b>Top Kategorien (Monat)</b>",
    "title_top_items":"🏷 <b>Top Artikel (Monat)</b>",
    "tmpl_none":"Noch keine Vorlagen.\n«➕ Neue Vorlage» drücken.",
    "tmpl_choose":"Vorlage oder Aktion wählen:",
    "tmpl_add_name":"Vorlagenname eingeben (z.B. Bus, Kaffee):",
    "tmpl_add_amount":"Betrag für «{name}» ({sym}):",
    "tmpl_add_cat":"Kategorie für «{name}»:",
    "tmpl_added":"✅ Vorlage «{name}» gespeichert ({amount:.2f} {sym}, {cat})",
    "tmpl_del_choose":"Welche Vorlage löschen?","tmpl_deleted":"🗑 Vorlage «{name}» gelöscht.",
    "tmpl_no_del":"Keine Vorlagen zum Löschen.",
    "lang_select":"🌐 Sprache wählen:",
    "curr_primary":"<b>Hauptwährung</b> wählen\n(Ausgaben werden darin eingegeben):",
    "curr_secondary":"<b>Umrechnungswährung</b> wählen\n(neben Beträgen angezeigt):",
    "curr_none":"🚫 Keine Umrechnung",
    "curr_set_primary":"✅ Hauptwährung: <b>{cur}</b> {sym}",
    "curr_set_secondary":"✅ Umrechnung: <b>{cur}</b> {sym}",
    "curr_set_secondary_none":"✅ Umrechnung deaktiviert.",
    "convert_prompt":"Betrag und Währungen in einer Nachricht senden:\n\n<code>12000 UAH USD</code>\n<code>100 EUR PLN</code>\n<code>50 GBP UAH</code>\n\nAlle Weltwährungscodes werden unterstützt.",
    "convert_result":"💱 <b>{amount:.2f} {from_cur}</b> = <b>{result:.2f} {to_cur}</b>\n📈 Kurs: 1 {from_cur} = {rate:.4f} {to_cur}",
    "convert_error":"❌ Umrechnung fehlgeschlagen.\n\nFormat: <code>12000 UAH USD</code>\nWährungscodes prüfen (3 lateinische Buchstaben).",
    "convert_unavail":"❌ Kurs für <b>{cur}</b> nicht verfügbar.",
    "export_period":"Zeitraum für Excel-Bericht wählen:",
    "export_btn_today":"📅 Heute","export_btn_week":"📆 Diese Woche",
    "export_btn_month":"🗓 Diesen Monat","export_btn_all":"📁 Alle Einträge",
    "export_empty":"Keine Daten für den Bericht.",
    "export_sending":"⏳ Datei wird erstellt...",
    "export_no_xlsx":"openpyxl installieren:\n<code>pip install openpyxl</code>",
    "export_sheet_all":"Ausgaben","export_sheet_cat":"Nach Kategorie","export_sheet_day":"Nach Tag",
    "ccat_header":"⚙️ <b>Meine Kategorien</b>",
    "ccat_list_item":"• {label}",
    "ccat_none":"Noch keine eigenen Kategorien.",
    "ccat_action":"Was möchtest du tun?",
    "ccat_add_name":"Namen der neuen Kategorie eingeben.\nEmoji möglich: <code>🚕 Taxi</code>",
    "ccat_too_long":"Name zu lang (max. 40 Zeichen).",
    "ccat_added":"✅ Kategorie «{name}» hinzugefügt!",
    "ccat_del_choose":"Welche Kategorie löschen?",
    "ccat_deleted":"🗑 Kategorie «{name}» gelöscht.",
    "ccat_no_del":"Keine eigenen Kategorien zum Löschen.",
    "btn_recurring":"🔁 Wiederkehrend","btn_reminders":"🔔 Erinnerungen","btn_compare":"📊 Vergleich",
    "recur_none":"Keine wiederkehrenden Ausgaben.","recur_choose":"Wählen oder hinzufügen:",
    "btn_recur_add":"➕ Hinzufügen","btn_recur_del":"🗑 Löschen",
    "recur_add_name":"Name (Miete, Netflix):","recur_add_amount":"Betrag für «{name}» ({sym}):",
    "recur_add_cat":"Kategorie für «{name}»:","recur_add_day":"Tag des Monats (1–31)\n(im kürzeren Monat am letzten Tag):",
    "recur_bad_day":"Zahl von 1 bis 28 eingeben. (1–31)",
    "recur_added":"✅ «{name}» hinzugefügt!\nJeden {day}. des Monats: {amount:.2f} {sym}",
    "recur_del_choose":"Welche Ausgabe löschen?","recur_deleted":"🗑 «{name}» gelöscht.","recur_no_del":"Nichts zu löschen.",
    "recur_auto":"🔁 Auto-Buchung: <b>{name}</b> — {amount:.2f} {sym}",
    "remind_title":"🔔 <b>Erinnerungen</b>",
    "remind_current":"Aktuell:\n• Inaktivität: {inact} Tage\n• Täglich: {daily}\n• Wöchentlich: {weekly} (Tag {wday})",
    "remind_off":"aus","remind_btn_inactive":"⏰ Inaktivität","remind_btn_daily":"📅 Täglich","remind_btn_weekly":"📆 Wöchentlich",
    "remind_inactive_prompt":"Nach wie vielen inaktiven Tagen erinnern?\n(0 = aus, max 30):",
    "remind_daily_prompt":"Tägliche Erinnerungszeit (HH:MM UTC) oder 0:","remind_weekly_day_prompt":"Wochentag (0=Mo..6=So) oder -1:",
    "remind_weekly_time_prompt":"Wöchentliche Erinnerungszeit (HH:MM UTC):",
    "remind_saved":"✅ Erinnerung gespeichert.","remind_bad_time":"Format: HH:MM (z.B. 09:00)",
    "remind_bad_inactive":"Zahl von 0 bis 30.","remind_bad_wday":"Zahl von -1 bis 6.",
    "remind_inact_msg":"👋 Du hast seit <b>{days} Tagen</b> keine Ausgaben eingetragen!",
    "remind_daily_msg":"📅 Zeit für deine Ausgaben! Drücke «{btn}».",
    "remind_weekly_msg":"📆 Woche endet! Drücke «{btn}».",
    "compare_title":"📊 <b>Monatsvergleich</b>","compare_this":"Dieser Monat","compare_prev":"Letzter Monat",
    "compare_diff_more":"▲ +{diff:.2f} {sym} (+{pct:.0f}%) mehr","compare_diff_less":"▼ {diff:.2f} {sym} ({pct:.0f}%) weniger",
    "compare_diff_same":"◆ Keine Änderung","compare_no_data":"Nicht genug Daten.","compare_by_cat":"Nach Kategorien:",
    "donate_msg":"☕ <b>Projekt unterstützen</b>\n\nWenn der Bot nützlich ist — über einen Kaffee würde ich mich freuen!\n\n{url}",
    "donate_no_url":"Spendenlink ist nicht konfiguriert.",
},
}

# ── Курси валют (open.er-api.com — 160+ валют, без ключа) ─────────────────────
_rates_cache: dict = {"rates_usd": {}, "updated": None}

def _fetch_rates_sync() -> dict:
    """Returns {currency_code: rate_per_usd}"""
    try:
        req = urllib.request.Request(
            "https://open.er-api.com/v6/latest/USD",
            headers={"User-Agent": "ExpenseBot/3.0"}
        )
        with urllib.request.urlopen(req, timeout=6) as r:
            data = json.loads(r.read())
        if data.get("result") == "success":
            return data.get("rates", {})
    except Exception:
        pass
    return {}

async def get_rates() -> dict:
    now = datetime.now()
    if (_rates_cache["rates_usd"] and _rates_cache["updated"] and
            (now - _rates_cache["updated"]).total_seconds() < 3600):
        return _rates_cache["rates_usd"]
    loop = asyncio.get_event_loop()
    rates = await loop.run_in_executor(None, _fetch_rates_sync)
    if rates:
        _rates_cache.update({"rates_usd": rates, "updated": now})
    return _rates_cache["rates_usd"]

def convert_amount(amount, from_cur: str, to_cur: str, rates_usd: dict) -> float | None:
    amount = float(amount)
    """Convert via USD as base. rates_usd = {code: per_usd}"""
    if from_cur == to_cur:
        return amount
    if from_cur not in rates_usd or to_cur not in rates_usd:
        return None
    return amount * rates_usd[to_cur] / rates_usd[from_cur]

def secondary_str(amount, s: dict, rates_usd: dict) -> str:
    amount = float(amount)
    primary   = s.get("primary_currency", "UAH")
    secondary = s.get("secondary_currency", "USD")
    if not secondary or secondary in ("none", primary):
        return ""
    conv = convert_amount(amount, primary, secondary, rates_usd)
    if conv is None:
        return ""
    s2 = CURRENCY_SYMBOLS.get(secondary, secondary)
    return f" (~{s2}{conv:.2f})"

# ── База даних ─────────────────────────────────────────────────────────────────
class _PgDB:
    """Wrapper for psycopg2 — PostgreSQL (Railway production)."""
    def __init__(self, conn):
        self._conn = conn
        self._cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)  # type: ignore

    def execute(self, sql: str, params=()):
        self._cur.execute(sql.replace("?", "%s"), params or ())
        return self._cur

    def __enter__(self):
        return self

    def __exit__(self, exc_type, *_):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._cur.close()
        self._conn.close()


class _SqliteDB:
    """Wrapper for sqlite3 — local development."""
    def __init__(self, conn):
        self._conn = conn
        self._conn.row_factory = sqlite3.Row
        self._cur = self._conn.cursor()

    def execute(self, sql: str, params=()):
        # PG-style %s → sqlite ?  and  SERIAL PRIMARY KEY → INTEGER PRIMARY KEY AUTOINCREMENT
        sql = sql.replace("%s", "?")
        self._cur.execute(sql, params or ())
        return self._cur

    def __enter__(self):
        return self

    def __exit__(self, exc_type, *_):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._conn.close()


def get_db():
    if USE_PG:
        if not PSYCOPG2_OK or psycopg2 is None:
            raise RuntimeError(
                "DATABASE_URL is set but psycopg2 is not installed!\n"
                "Add 'psycopg2-binary==2.9.9' to requirements.txt"
            )
        conn = psycopg2.connect(DATABASE_URL)
        return _PgDB(conn)
    else:
        d = os.path.dirname(DB_PATH)
        if d:
            os.makedirs(d, exist_ok=True)
        return _SqliteDB(sqlite3.connect(DB_PATH))


def init_db():
    """Create all tables. Uses PostgreSQL syntax on Railway, SQLite locally."""
    if USE_PG:
        statements = [
            """CREATE TABLE IF NOT EXISTS expenses (
                id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
                amount NUMERIC(12,2) NOT NULL, category TEXT NOT NULL,
                item_name TEXT NOT NULL, created TIMESTAMP NOT NULL DEFAULT NOW()
            )""",
            """CREATE TABLE IF NOT EXISTS limits (
                user_id BIGINT NOT NULL, category TEXT NOT NULL,
                amount NUMERIC(12,2) NOT NULL, PRIMARY KEY (user_id, category)
            )""",
            """CREATE TABLE IF NOT EXISTS user_settings (
                user_id BIGINT PRIMARY KEY, language TEXT NOT NULL DEFAULT 'uk',
                primary_currency TEXT NOT NULL DEFAULT 'UAH',
                secondary_currency TEXT DEFAULT 'USD'
            )""",
            """CREATE TABLE IF NOT EXISTS templates (
                id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
                name TEXT NOT NULL, amount NUMERIC(12,2) NOT NULL, category TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS custom_categories (
                id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL, label TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS recurring (
                id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
                name TEXT NOT NULL, amount NUMERIC(12,2) NOT NULL,
                category TEXT NOT NULL, day_of_month INTEGER NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS reminder_settings (
                user_id BIGINT PRIMARY KEY, inactive_days INTEGER NOT NULL DEFAULT 0,
                daily_time TEXT DEFAULT NULL, weekly_day INTEGER DEFAULT NULL,
                weekly_time TEXT DEFAULT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS income (
                id SERIAL PRIMARY KEY, user_id BIGINT NOT NULL,
                amount NUMERIC(12,2) NOT NULL, source TEXT NOT NULL,
                created TIMESTAMP NOT NULL DEFAULT NOW()
            )""",
        ]
        with get_db() as db:
            for stmt in statements:
                db.execute(stmt)
    else:
        # SQLite — uses executescript (single connection, no %s issue)
        conn = sqlite3.connect(DB_PATH)
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                amount REAL NOT NULL, category TEXT NOT NULL,
                item_name TEXT NOT NULL,
                created TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS limits (
                user_id INTEGER NOT NULL, category TEXT NOT NULL, amount REAL NOT NULL,
                PRIMARY KEY (user_id, category)
            );
            CREATE TABLE IF NOT EXISTS user_settings (
                user_id INTEGER PRIMARY KEY, language TEXT NOT NULL DEFAULT 'uk',
                primary_currency TEXT NOT NULL DEFAULT 'UAH',
                secondary_currency TEXT DEFAULT 'USD'
            );
            CREATE TABLE IF NOT EXISTS templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                name TEXT NOT NULL, amount REAL NOT NULL, category TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS custom_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL, label TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS recurring (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                name TEXT NOT NULL, amount REAL NOT NULL,
                category TEXT NOT NULL, day_of_month INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS reminder_settings (
                user_id INTEGER PRIMARY KEY,
                inactive_days INTEGER NOT NULL DEFAULT 0,
                daily_time TEXT DEFAULT NULL,
                weekly_day INTEGER DEFAULT NULL,
                weekly_time TEXT DEFAULT NULL
            );
            CREATE TABLE IF NOT EXISTS income (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                amount REAL NOT NULL, source TEXT NOT NULL,
                created TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );
        """)
        conn.commit()
        conn.close()
        logging.info("SQLite DB ready at %s", DB_PATH)

# ── Налаштування ───────────────────────────────────────────────────────────────
_DS = {"language":"uk","primary_currency":"UAH","secondary_currency":"USD"}

# Telegram locale → bot language
TG_LANG_MAP = {
    "uk": "uk", "ru": "ru",
    "en": "en", "en-us": "en", "en-gb": "en",
    "de": "de", "de-de": "de", "de-at": "de", "de-ch": "de",
}

def detect_lang(tg_lang_code: str | None) -> str:
    if not tg_lang_code:
        return "uk"
    code = tg_lang_code.lower()
    if code in TG_LANG_MAP:
        return TG_LANG_MAP[code]
    return TG_LANG_MAP.get(code.split("-")[0], "uk")

def get_settings(uid: int) -> dict:
    with get_db() as db:
        row = db.execute("SELECT * FROM user_settings WHERE user_id=?", (uid,)).fetchone()
    return dict(row) if row else dict(_DS)

def is_first_time(uid: int) -> bool:
    with get_db() as db:
        row = db.execute("SELECT 1 FROM user_settings WHERE user_id=?", (uid,)).fetchone()
    return row is None

def save_settings(uid: int, **kw):
    s = get_settings(uid)
    s.update(kw)
    with get_db() as db:
        db.execute(
            "INSERT INTO user_settings "
            "(user_id,language,primary_currency,secondary_currency) VALUES (%s,%s,%s,%s) "
            "ON CONFLICT (user_id) DO UPDATE SET language=EXCLUDED.language,"
            "primary_currency=EXCLUDED.primary_currency,secondary_currency=EXCLUDED.secondary_currency",
            (uid, s["language"], s["primary_currency"], s.get("secondary_currency","USD"))
        )

# ── Переклад ───────────────────────────────────────────────────────────────────
def tr(uid: int, key: str, s: dict | None = None, **kw) -> str:
    lang = (s or get_settings(uid)).get("language","uk")
    text = T.get(lang, T["uk"]).get(key, T["uk"].get(key, key))
    return text.format(**kw) if kw else text

def sym(s: dict) -> str:
    return CURRENCY_SYMBOLS.get(s.get("primary_currency","UAH"), "₴")

# ── Категорії (вбудовані + власні) ────────────────────────────────────────────
def get_custom_cats(uid: int) -> list[dict]:
    with get_db() as db:
        return [dict(r) for r in db.execute(
            "SELECT * FROM custom_categories WHERE user_id=? ORDER BY label", (uid,)
        ).fetchall()]

def cat_label(lang: str, key: str, uid: int | None = None) -> str:
    if key.startswith("cust:") and uid is not None:
        cid = int(key.split(":")[1])
        with get_db() as db:
            row = db.execute("SELECT label FROM custom_categories WHERE id=?", (cid,)).fetchone()
        return row["label"] if row else key
    return CATEGORY_LABELS.get(lang, CATEGORY_LABELS["uk"]).get(key, key)

def cat_key_from_label(label: str, lang: str, uid: int) -> str | None:
    for k, v in CATEGORY_LABELS.get(lang, {}).items():
        if v == label:
            return k
    for c in get_custom_cats(uid):
        if c["label"] == label:
            return f"cust:{c['id']}"
    return None

# ── Income DB ────────────────────────────────────────────────────────────────
def add_income(uid: int, amount: float, source: str):
    with get_db() as db:
        db.execute(
            "INSERT INTO income (user_id, amount, source) VALUES (?, ?, ?)",
            (uid, amount, source)
        )

def get_month_income(uid: int) -> float:
    start, end = period_dates("month")
    with get_db() as db:
        row = db.execute(
            "SELECT COALESCE(SUM(amount), 0) as v FROM income "
            "WHERE user_id=? AND created BETWEEN ? AND ?",
            (uid, start, end)
        ).fetchone()
    return float(row["v"]) if row else 0.0

def get_month_expenses_total(uid: int) -> float:
    start, end = period_dates("month")
    with get_db() as db:
        row = db.execute(
            "SELECT COALESCE(SUM(amount), 0) as v FROM expenses "
            "WHERE user_id=? AND created BETWEEN ? AND ?",
            (uid, start, end)
        ).fetchone()
    return float(row["v"]) if row else 0.0

# ── Безпека ─────────────────────────────────────────────────────────────────────
def _is_allowed(uid: int) -> bool:
    """Returns True when ALLOWED_USERS is empty (public) or uid is whitelisted."""
    return not ALLOWED_USERS or uid in ALLOWED_USERS

def sanitize(text: str, max_len: int = MAX_INPUT) -> str:
    return (text or "").strip()[:max_len]

def parse_amount(text: str) -> float | None:
    try:
        v = float((text or "").replace(",", ".").replace(" ", ""))
        return v if 0 < v <= MAX_AMOUNT else None
    except ValueError:
        return None

# ── Recurring DB ─────────────────────────────────────────────────────────────
def get_recurring(uid: int) -> list[dict]:
    with get_db() as db:
        return [dict(r) for r in db.execute(
            "SELECT * FROM recurring WHERE user_id=? ORDER BY day_of_month,name", (uid,)).fetchall()]

def get_all_recurring() -> list[dict]:
    with get_db() as db:
        return [dict(r) for r in db.execute("SELECT * FROM recurring").fetchall()]

def add_recurring_db(uid: int, name: str, amount: float, category: str, day: int):
    with get_db() as db:
        db.execute(
            "INSERT INTO recurring (user_id,name,amount,category,day_of_month) VALUES (?,?,?,?,?)",
            (uid, name, amount, category, day))

def del_recurring_db(rec_id: int):
    with get_db() as db:
        db.execute("DELETE FROM recurring WHERE id=?", (rec_id,))

# ── Reminder DB ──────────────────────────────────────────────────────────────
def get_reminder_settings(uid: int) -> dict:
    with get_db() as db:
        row = db.execute("SELECT * FROM reminder_settings WHERE user_id=?", (uid,)).fetchone()
    return dict(row) if row else {
        "user_id": uid, "inactive_days": 0,
        "daily_time": None, "weekly_day": None, "weekly_time": None
    }

def save_reminder_settings(uid: int, **kw):
    rs = get_reminder_settings(uid); rs.update(kw)
    with get_db() as db:
        db.execute(
            "INSERT INTO reminder_settings "
            "(user_id,inactive_days,daily_time,weekly_day,weekly_time) VALUES (%s,%s,%s,%s,%s) "
            "ON CONFLICT (user_id) DO UPDATE SET inactive_days=EXCLUDED.inactive_days, "
            "daily_time=EXCLUDED.daily_time, weekly_day=EXCLUDED.weekly_day, "
            "weekly_time=EXCLUDED.weekly_time",
            (uid, rs["inactive_days"], rs["daily_time"], rs["weekly_day"], rs["weekly_time"])
        )

def get_all_users_with_reminders() -> list[dict]:
    with get_db() as db:
        return [dict(r) for r in db.execute(
            "SELECT * FROM reminder_settings WHERE inactive_days>0 OR daily_time IS NOT NULL "
            "OR weekly_day IS NOT NULL"
        ).fetchall()]

def get_last_expense_date(uid: int) -> str | None:
    with get_db() as db:
        row = db.execute(
            "SELECT MAX(created) as last FROM expenses WHERE user_id=?", (uid,)
        ).fetchone()
    return row["last"] if row and row["last"] else None

# ── Клавіатури ─────────────────────────────────────────────────────────────────

def main_kb(uid: int, s: dict | None = None) -> ReplyKeyboardMarkup:
    g = lambda k: tr(uid, k, s)
    return ReplyKeyboardMarkup([
        [g("btn_add"),      g("btn_quick")],
        [g("btn_finance"),  g("btn_reports")],
        [g("btn_more"),     g("btn_settings")],
    ], resize_keyboard=True)

def finance_kb(uid: int, s: dict | None = None) -> ReplyKeyboardMarkup:
    g = lambda k: tr(uid, k, s)
    return ReplyKeyboardMarkup([
        [g("btn_add_income_short")],
        [g("btn_back")],
    ], resize_keyboard=True)

def reports_kb(uid: int, s: dict | None = None) -> ReplyKeyboardMarkup:
    g = lambda k: tr(uid, k, s)
    return ReplyKeyboardMarkup([
        [g("btn_today"),    g("btn_week")],
        [g("btn_month"),    g("btn_compare")],
        [g("btn_top_cat"),  g("btn_top_items")],
        [g("btn_export")],
        [g("btn_back")],
    ], resize_keyboard=True)

def more_kb(uid: int, s: dict | None = None) -> ReplyKeyboardMarkup:
    g = lambda k: tr(uid, k, s)
    return ReplyKeyboardMarkup([
        [g("btn_limit"),    g("btn_recurring")],
        [g("btn_reminders"),g("btn_convert")],
        [g("btn_my_cats"),  g("btn_help")],
        [g("btn_feedback")],
        [g("btn_back")],
    ], resize_keyboard=True)

def cat_kb(uid: int, s: dict) -> ReplyKeyboardMarkup:
    lang = s["language"]
    btns = [[CATEGORY_LABELS[lang][k]] for k in BUILT_IN_KEYS]
    for c in get_custom_cats(uid):
        btns.append([c["label"]])
    btns.append([tr(uid, "btn_cancel", s)])
    return ReplyKeyboardMarkup(btns, resize_keyboard=True, one_time_keyboard=True)

def cancel_kb(uid: int, s: dict) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([[tr(uid, "btn_cancel", s)]], resize_keyboard=True)

def lang_kb(uid: int, s: dict) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[v] for v in LANG_BUTTONS.values()] + [[tr(uid, "btn_cancel", s)]],
        resize_keyboard=True, one_time_keyboard=True
    )

def curr_kb(uid: int, s: dict, with_none=False) -> ReplyKeyboardMarkup:
    btns = [[v] for v in CURRENCY_BUTTONS.values()]
    if with_none:
        btns.append([tr(uid, "curr_none", s)])
    btns.append([tr(uid, "btn_cancel", s)])
    return ReplyKeyboardMarkup(btns, resize_keyboard=True, one_time_keyboard=True)

def tmpl_kb(uid: int, s: dict, templates: list) -> ReplyKeyboardMarkup:
    _sym = sym(s)
    btns = [[f"⚡ {t['name']} ({t['amount']:.0f} {_sym})"] for t in templates]
    btns += [[tr(uid, "btn_tmpl_add", s), tr(uid, "btn_tmpl_del", s)],
             [tr(uid, "btn_cancel", s)]]
    return ReplyKeyboardMarkup(btns, resize_keyboard=True, one_time_keyboard=True)

def export_kb(uid: int, s: dict) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [tr(uid, "export_btn_today", s), tr(uid, "export_btn_week", s)],
        [tr(uid, "export_btn_month", s), tr(uid, "export_btn_all", s)],
        [tr(uid, "btn_cancel", s)],
    ], resize_keyboard=True, one_time_keyboard=True)

def settings_kb(uid: int, s: dict) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [tr(uid, 'btn_lang', s),     tr(uid, 'btn_currency', s)],
        [tr(uid, 'btn_donate', s)],
        [tr(uid, 'btn_cancel', s)],
    ], resize_keyboard=True, one_time_keyboard=True)

def recur_kb(uid: int, s: dict, items: list) -> ReplyKeyboardMarkup:
    _sym = sym(s)
    rows = [[f"🔁 {r['name']} — {r['amount']:.0f} {_sym} ({r['day_of_month']}-го)"] for r in items]
    rows.append([tr(uid, "btn_recur_add", s), tr(uid, "btn_recur_del", s)])
    rows.append([tr(uid, "btn_cancel", s)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True)

def recur_del_kb(uid: int, s: dict, items: list) -> ReplyKeyboardMarkup:
    rows = [[f"🗑 {r['name']}"] for r in items]
    rows.append([tr(uid, "btn_cancel", s)])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True)

def remind_kb(uid: int, s: dict) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([
        [tr(uid, "remind_btn_inactive", s)],
        [tr(uid, "remind_btn_daily", s)],
        [tr(uid, "remind_btn_weekly", s)],
        [tr(uid, "btn_cancel", s)],
    ], resize_keyboard=True, one_time_keyboard=True)

# ── Допоміжні ──────────────────────────────────────────────────────────────────
def period_dates(period: str):
    now = datetime.now()
    if period == "day":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = (now - timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0)
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:  # all
        start = datetime(2000, 1, 1)
    fmt = "%Y-%m-%d %H:%M:%S"
    return start.strftime(fmt), now.strftime(fmt)

def check_limit(uid: int, cat_key: str, s: dict) -> str | None:
    _sym = sym(s)
    with get_db() as db:
        row = db.execute("SELECT amount FROM limits WHERE user_id=? AND category=?",
                         (uid, cat_key)).fetchone()
        if not row:
            return None
        limit = row["amount"]
        start, end = period_dates("month")
        spent = db.execute(
            "SELECT COALESCE(SUM(amount),0) as v FROM expenses "
            "WHERE user_id=? AND category=? AND created BETWEEN ? AND ?",
            (uid, cat_key, start, end)
        ).fetchone()["v"]
    cat = cat_label(s["language"], cat_key, uid)
    if spent >= limit:
        return tr(uid, "limit_over", s, cat=cat, spent=spent, limit=limit, sym=_sym)
    if spent >= limit * 0.8:
        return tr(uid, "limit_warn", s, cat=cat, pct=spent/limit*100,
                  spent=spent, limit=limit, sym=_sym)
    return None

def fmt_date(date_str: str, lang: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return f"{dt.day:02d} {MONTHS[lang][dt.month]}"

def fmt_today(uid: int, rows, s: dict, rates: dict) -> str:
    title = tr(uid, "title_today", s)
    if not rows:
        return f"{title}\n\n{tr(uid, 'no_data', s)}"
    total = sum(float(r["amount"]) for r in rows)
    _sym, sec = sym(s), secondary_str(total, s, rates)
    lines = [title, f"{tr(uid, 'total_label', s)}: <b>{total:.2f} {_sym}{sec}</b>\n"]
    by_cat: dict[str, float] = defaultdict(float)
    for r in rows:
        by_cat[r["category"]] += r["amount"]
    for k, amt in sorted(by_cat.items(), key=lambda x: -x[1]):
        lines.append(f"  {cat_label(s['language'], k, uid)}: {amt:.2f} {_sym}")
    return "\n".join(lines)

def fmt_period(uid: int, rows, title_key: str, s: dict, rates: dict) -> str:
    title = tr(uid, title_key, s)
    if not rows:
        return f"{title}\n\n{tr(uid, 'no_data', s)}"
    lang, _sym = s["language"], sym(s)
    total = sum(float(r["amount"]) for r in rows)
    sec = secondary_str(total, s, rates)
    lines = [title, f"{tr(uid, 'total_label', s)}: <b>{total:.2f} {_sym}{sec}</b>"]
    by_date: dict[str, dict] = defaultdict(lambda: defaultdict(float))
    for r in rows:
        by_date[str(r["created"])[:10]][r["category"]] += r["amount"]
    for date in sorted(by_date):
        day_total = sum(by_date[date].values())
        day_sec = secondary_str(day_total, s, rates)
        lines.append(f"\n<b>📅 {fmt_date(date, lang)}</b>  —  {day_total:.2f} {_sym}{day_sec}")
        for k, amt in sorted(by_date[date].items(), key=lambda x: -x[1]):
            lines.append(f"   {cat_label(lang, k, uid)}: {amt:.2f} {_sym}")
    return "\n".join(lines)

# ── Excel генерація ────────────────────────────────────────────────────────────
XLSX_I18N = {
    "uk": {
        "num": "№", "date": "Дата", "time": "Час", "category": "Категорія",
        "item": "Товар / Послуга", "amount": "Сума", "percent": "%",
        "total": "РАЗОМ", "avg_day": "Середнє за день",
        "days": "Днів з витратами", "operations": "Операцій",
        "period_from": "Період з", "period_to": "по",
        "generated": "Сформовано",
        "summary_title": "Зведення",
    },
    "ru": {
        "num": "№", "date": "Дата", "time": "Время", "category": "Категория",
        "item": "Товар / Услуга", "amount": "Сумма", "percent": "%",
        "total": "ИТОГО", "avg_day": "Среднее в день",
        "days": "Дней с расходами", "operations": "Операций",
        "period_from": "Период с", "period_to": "по",
        "generated": "Сформировано",
        "summary_title": "Сводка",
    },
    "en": {
        "num": "#", "date": "Date", "time": "Time", "category": "Category",
        "item": "Item / Service", "amount": "Amount", "percent": "%",
        "total": "TOTAL", "avg_day": "Daily average",
        "days": "Days with expenses", "operations": "Operations",
        "period_from": "Period from", "period_to": "to",
        "generated": "Generated",
        "summary_title": "Summary",
    },
    "de": {
        "num": "#", "date": "Datum", "time": "Uhrzeit", "category": "Kategorie",
        "item": "Artikel / Dienstleistung", "amount": "Betrag", "percent": "%",
        "total": "GESAMT", "avg_day": "Tagesdurchschnitt",
        "days": "Tage mit Ausgaben", "operations": "Buchungen",
        "period_from": "Zeitraum von", "period_to": "bis",
        "generated": "Erstellt",
        "summary_title": "Übersicht",
    },
}

def build_xlsx(uid: int, rows: list, s: dict, period_label: str = "") -> bytes:
    wb = openpyxl.Workbook()
    lang = s.get("language", "uk")
    cur  = s.get("primary_currency", "UAH")
    _sym = sym(s)
    i18 = XLSX_I18N.get(lang, XLSX_I18N["en"])

    # ── Стилі ──────────────────────────────────────────────────────────────────
    GREEN       = "2D6A4F"
    GREEN_LIGHT = "EAF4EF"
    AMBER       = "E9A319"
    GRAY_HDR    = "F2F2F2"

    def hfont(color="FFFFFF", bold=True, size=11):
        return Font(bold=bold, color=color, size=size)

    def fill(color): return PatternFill("solid", fgColor=color)

    thin   = Side(style="thin",   color="BBBBBB")
    medium = Side(style="medium", color="888888")
    def bord(top=thin, bot=thin, lft=thin, rgt=thin):
        return Border(top=top, bottom=bot, left=lft, right=rgt)

    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    num_fmt = f'#,##0.00 "{_sym}"'

    def hdr_row(ws, row_idx):
        for cell in ws[row_idx]:
            cell.font      = hfont()
            cell.fill      = fill(GREEN)
            cell.alignment = center
            cell.border    = bord(top=Side(style="medium", color=GREEN),
                                  bot=Side(style="medium", color=GREEN))

    def total_row(ws, row_idx):
        for cell in ws[row_idx]:
            cell.font   = hfont(color="000000", bold=True)
            cell.fill   = fill(GREEN_LIGHT)
            cell.border = bord(top=Side(style="medium", color=GREEN))

    def alt_row(ws, row_idx, is_alt: bool):
        for cell in ws[row_idx]:
            if is_alt:
                cell.fill = fill("F7FCF9")
            cell.border = bord()

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 0 — ЗВЕДЕННЯ
    # ══════════════════════════════════════════════════════════════════════════
    ws0 = wb.active
    ws0.title = i18["summary_title"]
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 22

    total       = sum(r["amount"] for r in rows)
    by_cat: dict[str, float] = defaultdict(float)
    by_day: dict[str, float] = defaultdict(float)
    for r in rows:
        by_cat[r["category"]] += r["amount"]
        by_day[str(r["created"])[:10]] += r["amount"]
    n_days = len(by_day)
    n_ops  = len(rows)
    avg    = total / n_days if n_days else 0
    date_from = min(by_day) if by_day else "—"
    date_to   = max(by_day) if by_day else "—"

    # Заголовок блоку
    def kv(ws, key, val, key_bold=False, val_fmt=None, val_num=False, shade=False):
        ws.append([key, val])
        ri = ws.max_row
        kc, vc = ws.cell(ri, 1), ws.cell(ri, 2)
        kc.alignment = left
        vc.alignment = right
        kc.border = bord()
        vc.border = bord()
        if key_bold:
            kc.font = hfont(color="000000")
        if shade:
            kc.fill = fill(GRAY_HDR)
            vc.fill = fill(GRAY_HDR)
        if val_num:
            vc.number_format = num_fmt
        if val_fmt:
            vc.number_format = val_fmt

    # Шапка зведення
    ws0.append([i18["summary_title"], ""])
    ws0.merge_cells("A1:B1")
    title_cell = ws0["A1"]
    title_cell.value    = f'📊  {i18["summary_title"]}  —  {period_label}'
    title_cell.font     = hfont(size=13)
    title_cell.fill     = fill(GREEN)
    title_cell.alignment= center
    ws0.row_dimensions[1].height = 26

    ws0.append(["", ""])  # пустий рядок

    kv(ws0, i18["period_from"], date_from, shade=True)
    kv(ws0, i18["period_to"],   date_to,   shade=True)
    kv(ws0, i18["operations"],  n_ops,     shade=False)
    kv(ws0, i18["days"],        n_days,    shade=False)

    ws0.append(["", ""])

    # Підсумок
    kv(ws0, i18["total"],    total, key_bold=True, val_num=True, shade=True)
    kv(ws0, i18["avg_day"],  avg,   key_bold=False, val_num=True, shade=True)

    ws0.append(["", ""])

    # По категоріях у зведенні
    ws0.append([i18["category"], i18["amount"]])
    hdr_row(ws0, ws0.max_row)
    for i, (k, amt) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]), 1):
        pct = amt / total * 100 if total else 0
        label = cat_label(lang, k, uid)
        ws0.append([f"{label}  ({pct:.1f}%)", amt])
        ri = ws0.max_row
        ws0.cell(ri, 1).alignment = left
        ws0.cell(ri, 1).border    = bord()
        ws0.cell(ri, 2).number_format = num_fmt
        ws0.cell(ri, 2).alignment = right
        ws0.cell(ri, 2).border    = bord()
        if i % 2 == 0:
            ws0.cell(ri, 1).fill = fill("F7FCF9")
            ws0.cell(ri, 2).fill = fill("F7FCF9")

    # Підвал з датою генерації
    ws0.append(["", ""])
    ws0.append([i18["generated"], datetime.now().strftime("%Y-%m-%d %H:%M")])
    last = ws0.max_row
    ws0.cell(last, 1).font = Font(italic=True, color="999999", size=9)
    ws0.cell(last, 2).font = Font(italic=True, color="999999", size=9)

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 1 — ВСІ ВИТРАТИ
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet(tr(uid, "export_sheet_all", s))
    hdrs = [i18["num"], i18["date"], i18["time"], i18["category"], i18["item"],
            f'{i18["amount"]} ({cur})']
    ws1.append(hdrs)
    hdr_row(ws1, 1)
    ws1.row_dimensions[1].height = 20
    for col, w in zip("ABCDEF", [5, 13, 8, 22, 32, 16]):
        ws1.column_dimensions[col].width = w

    for i, r in enumerate(rows, 1):
        dt        = r["created"]
        date_part = dt[:10]
        time_part = dt[11:16] if len(dt) >= 16 else ""
        clabel    = cat_label(lang, r["category"], uid)
        ws1.append([i, date_part, time_part, clabel, r["item_name"], r["amount"]])
        ri = ws1.max_row
        alt_row(ws1, ri, i % 2 == 0)
        ws1.cell(ri, 1).alignment = center
        ws1.cell(ri, 2).alignment = center
        ws1.cell(ri, 3).alignment = center
        ws1.cell(ri, 6).number_format = num_fmt
        ws1.cell(ri, 6).alignment = right

    # Рядок підсумку
    ws1.append(["", "", "", "", i18["total"], total])
    total_row(ws1, ws1.max_row)
    ws1.cell(ws1.max_row, 5).alignment = right
    ws1.cell(ws1.max_row, 5).font      = Font(bold=True)
    ws1.cell(ws1.max_row, 6).number_format = num_fmt
    ws1.cell(ws1.max_row, 6).alignment = right

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 2 — ПО КАТЕГОРІЯХ
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(tr(uid, "export_sheet_cat", s))
    ws2.append([i18["category"], f'{i18["amount"]} ({cur})', i18["percent"]])
    hdr_row(ws2, 1)
    ws2.row_dimensions[1].height = 20
    for col, w in zip("ABC", [26, 18, 10]):
        ws2.column_dimensions[col].width = w

    for i, (k, amt) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]), 1):
        pct = round(amt / total * 100, 1) if total else 0
        ws2.append([cat_label(lang, k, uid), round(amt, 2), f"{pct}%"])
        ri = ws2.max_row
        alt_row(ws2, ri, i % 2 == 0)
        ws2.cell(ri, 2).number_format = num_fmt
        ws2.cell(ri, 2).alignment = right
        ws2.cell(ri, 3).alignment = center

    ws2.append([i18["total"], round(total, 2), "100%"])
    total_row(ws2, ws2.max_row)
    ws2.cell(ws2.max_row, 1).font      = Font(bold=True)
    ws2.cell(ws2.max_row, 2).number_format = num_fmt
    ws2.cell(ws2.max_row, 2).alignment = right
    ws2.cell(ws2.max_row, 3).alignment = center

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 3 — ПО ДНЯХ
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet(tr(uid, "export_sheet_day", s))
    ws3.append([i18["date"], f'{i18["amount"]} ({cur})'])
    hdr_row(ws3, 1)
    ws3.row_dimensions[1].height = 20
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 18

    for i, date in enumerate(sorted(by_day), 1):
        ws3.append([date, round(by_day[date], 2)])
        ri = ws3.max_row
        alt_row(ws3, ri, i % 2 == 0)
        ws3.cell(ri, 1).alignment = center
        ws3.cell(ri, 2).number_format = num_fmt
        ws3.cell(ri, 2).alignment = right

    ws3.append([i18["total"], round(total, 2)])
    total_row(ws3, ws3.max_row)
    ws3.cell(ws3.max_row, 1).font      = Font(bold=True)
    ws3.cell(ws3.max_row, 1).alignment = center
    ws3.cell(ws3.max_row, 2).number_format = num_fmt
    ws3.cell(ws3.max_row, 2).alignment = right

    ws3.append([i18["avg_day"], round(avg, 2)])
    ri = ws3.max_row
    ws3.cell(ri, 1).font           = Font(italic=True)
    ws3.cell(ri, 1).alignment      = center
    ws3.cell(ri, 2).number_format  = num_fmt
    ws3.cell(ri, 2).alignment      = right
    ws3.cell(ri, 2).border         = bord()
    ws3.cell(ri, 1).border         = bord()

    # Зведення — активний лист при відкритті
    wb.active = ws0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── /start ─────────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not _is_allowed(update.effective_user.id):
        return
    uid  = update.effective_user.id
    user = update.effective_user
    # Auto-detect language on first use (only if not already set by user)
    with get_db() as db:
        exists = db.execute(
            "SELECT 1 FROM user_settings WHERE user_id=?", (uid,)
        ).fetchone()
    if not exists:
        lang = detect_lang(user.language_code)
        save_settings(uid, language=lang)
    s = get_settings(uid)
    await update.message.reply_text(tr(uid, "welcome", s), parse_mode="HTML",
                                    reply_markup=main_kb(uid, s))

# ── Донат ───────────────────────────────────────────────────────────────────────
async def cmd_donate(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    if DONATE_URL:
        await update.message.reply_text(tr(uid, "donate_msg", s, url=DONATE_URL),
                                        parse_mode="HTML", reply_markup=main_kb(uid, s))
    else:
        await update.message.reply_text(tr(uid, "donate_no_url", s), reply_markup=main_kb(uid, s))

# ── Вибір мови ─────────────────────────────────────────────────────────────────
async def lang_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    await update.message.reply_text(tr(uid, "lang_select", s), reply_markup=lang_kb(uid, s))
    return LANG_SELECT

async def lang_select(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    chosen = next((k for k, v in LANG_BUTTONS.items() if v == text), None)
    if not chosen:
        await update.message.reply_text(tr(uid, "lang_select", s), reply_markup=lang_kb(uid, s))
        return LANG_SELECT
    save_settings(uid, language=chosen)
    new_s = get_settings(uid)
    await update.message.reply_text(T[chosen]["lang_switched"], parse_mode="HTML",
                                    reply_markup=main_kb(uid, new_s))
    return ConversationHandler.END

# ── Вибір валюти ───────────────────────────────────────────────────────────────
async def curr_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    await update.message.reply_text(tr(uid, "curr_primary", s), parse_mode="HTML",
                                    reply_markup=curr_kb(uid, s))
    return CURR_PRIMARY

async def curr_primary(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    chosen = next((k for k, v in CURRENCY_BUTTONS.items() if v == text), None)
    if not chosen:
        await update.message.reply_text(tr(uid, "curr_primary", s), parse_mode="HTML",
                                        reply_markup=curr_kb(uid, s))
        return CURR_PRIMARY
    save_settings(uid, primary_currency=chosen)
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "curr_set_primary", s, cur=chosen, sym=CURRENCY_SYMBOLS[chosen]) +
        f"\n\n{tr(uid, 'curr_secondary', s)}",
        parse_mode="HTML", reply_markup=curr_kb(uid, s, with_none=True)
    )
    return CURR_SECONDARY

async def curr_secondary(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    if text == tr(uid, "curr_none", s):
        save_settings(uid, secondary_currency="none")
        s = get_settings(uid)
        await update.message.reply_text(tr(uid, "curr_set_secondary_none", s),
                                        reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    chosen = next((k for k, v in CURRENCY_BUTTONS.items() if v == text), None)
    if not chosen:
        await update.message.reply_text(tr(uid, "curr_secondary", s), parse_mode="HTML",
                                        reply_markup=curr_kb(uid, s, with_none=True))
        return CURR_SECONDARY
    save_settings(uid, secondary_currency=chosen)
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "curr_set_secondary", s, cur=chosen, sym=CURRENCY_SYMBOLS[chosen]),
        parse_mode="HTML", reply_markup=main_kb(uid, s)
    )
    return ConversationHandler.END

# ── Конвертер ──────────────────────────────────────────────────────────────────
async def convert_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    await update.message.reply_text(tr(uid, "convert_prompt", s), parse_mode="HTML",
                                    reply_markup=cancel_kb(uid, s))
    return CONVERT

async def convert_do(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text.strip()
    s = get_settings(uid)

    # Exit if user pressed any main menu button (not just cancel)
    all_menu_btns = {T[l][k] for l in T for k in T[l]}
    if text in all_menu_btns and text != tr(uid, "btn_cancel", s):
        # Re-route to menu_router so the button works correctly
        await menu_router(update, ctx)
        return ConversationHandler.END

    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END

    m = re.match(r"([\d.,]+)\s+([A-Za-z]{3})\s+([A-Za-z]{3})", text)
    if not m:
        await update.message.reply_text(tr(uid, "convert_error", s), parse_mode="HTML",
                                        reply_markup=cancel_kb(uid, s))
        return CONVERT

    amount = float(m.group(1).replace(",", "."))
    from_cur, to_cur = m.group(2).upper(), m.group(3).upper()
    rates = await get_rates()

    if from_cur not in rates:
        await update.message.reply_text(tr(uid, "convert_unavail", s, cur=from_cur),
                                        parse_mode="HTML")
        return CONVERT
    if to_cur not in rates:
        await update.message.reply_text(tr(uid, "convert_unavail", s, cur=to_cur),
                                        parse_mode="HTML")
        return CONVERT

    result = convert_amount(amount, from_cur, to_cur, rates)
    rate_1  = convert_amount(1, from_cur, to_cur, rates)
    await update.message.reply_text(
        tr(uid, "convert_result", s, amount=amount, from_cur=from_cur,
           result=result, to_cur=to_cur, rate=rate_1),
        parse_mode="HTML", reply_markup=main_kb(uid, s)
    )
    return ConversationHandler.END

# ── Excel експорт ──────────────────────────────────────────────────────────────
async def export_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    if not XLSX_OK:
        await update.message.reply_text(tr(uid, "export_no_xlsx", s), parse_mode="HTML",
                                        reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    await update.message.reply_text(tr(uid, "export_period", s), reply_markup=export_kb(uid, s))
    return EXPORT_PERIOD

async def export_do(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)

    # Перевіряємо export-кнопки ПЕРШИМИ — до будь-якого escape
    period_map = {
        tr(uid, "export_btn_today", s): "day",
        tr(uid, "export_btn_week", s):  "week",
        tr(uid, "export_btn_month", s): "month",
        tr(uid, "export_btn_all", s):   "all",
    }
    period = period_map.get(text)

    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END

    if not period:
        # Тільки тепер перевіряємо чи це кнопка головного меню
        _main_btns = {T[_l][_k] for _l in T for _k in T[_l]
                      if _k.startswith("btn_") and not _k.startswith("btn_export")}
        if text in _main_btns:
            await menu_router(update, ctx)
            return ConversationHandler.END
        await update.message.reply_text(tr(uid, "export_period", s), reply_markup=export_kb(uid, s))
        return EXPORT_PERIOD

    start, end = period_dates(period)
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM expenses WHERE user_id=? AND created BETWEEN ? AND ? ORDER BY created",
            (uid, start, end)
        ).fetchall()

    if not rows:
        await update.message.reply_text(tr(uid, "export_empty", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END

    await update.message.reply_text(tr(uid, "export_sending", s), reply_markup=main_kb(uid, s))
    period_label = text  # кнопка яку натиснув користувач ("📅 Сьогодні" тощо)
    xlsx_bytes = build_xlsx(uid, rows, s, period_label=period_label)
    filename = f"expenses_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    await update.message.reply_document(
        document=io.BytesIO(xlsx_bytes),
        filename=filename,
    )
    return ConversationHandler.END

# ── Власні категорії ───────────────────────────────────────────────────────────
async def ccat_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    cats = get_custom_cats(uid)
    header = tr(uid, "ccat_header", s)
    if cats:
        items = "\n".join(tr(uid, "ccat_list_item", s, label=c["label"]) for c in cats)
        msg = f"{header}\n\n{items}\n\n{tr(uid, 'ccat_action', s)}"
    else:
        msg = f"{header}\n\n{tr(uid, 'ccat_none', s)}\n\n{tr(uid, 'ccat_action', s)}"
    btns = [
        [tr(uid, "btn_ccat_add", s), tr(uid, "btn_ccat_del", s)],
        [tr(uid, "btn_cancel", s)],
    ]
    await update.message.reply_text(msg, parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(btns, resize_keyboard=True, one_time_keyboard=True))
    return CCAT_MENU

async def ccat_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    if text == tr(uid, "btn_ccat_add", s):
        await update.message.reply_text(tr(uid, "ccat_add_name", s), parse_mode="HTML",
                                        reply_markup=cancel_kb(uid, s))
        return CCAT_NAME
    if text == tr(uid, "btn_ccat_del", s):
        cats = get_custom_cats(uid)
        if not cats:
            await update.message.reply_text(tr(uid, "ccat_no_del", s), reply_markup=main_kb(uid, s))
            return ConversationHandler.END
        ctx.user_data["ccat_del_map"] = {c["label"]: c["id"] for c in cats}
        btns = [[c["label"]] for c in cats] + [[tr(uid, "btn_cancel", s)]]
        await update.message.reply_text(tr(uid, "ccat_del_choose", s),
            reply_markup=ReplyKeyboardMarkup(btns, resize_keyboard=True, one_time_keyboard=True))
        return CCAT_DEL
    await update.message.reply_text(tr(uid, "ccat_action", s))
    return CCAT_MENU

async def ccat_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    label = text.strip()
    if len(label) > 40:
        await update.message.reply_text(tr(uid, "ccat_too_long", s))
        return CCAT_NAME
    with get_db() as db:
        db.execute("INSERT INTO custom_categories (user_id, label) VALUES (?,?)", (uid, label))
    await update.message.reply_text(tr(uid, "ccat_added", s, name=label), reply_markup=main_kb(uid, s))
    return ConversationHandler.END

async def ccat_del(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    cat_id = ctx.user_data.get("ccat_del_map", {}).get(text)
    if cat_id:
        with get_db() as db:
            db.execute("DELETE FROM custom_categories WHERE id=?", (cat_id,))
        await update.message.reply_text(tr(uid, "ccat_deleted", s, name=text),
                                        reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    await update.message.reply_text(tr(uid, "ccat_del_choose", s))
    return CCAT_DEL

# ── Додати витрату ─────────────────────────────────────────────────────────────
async def add_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    await update.message.reply_text(tr(uid, "choose_cat", s), reply_markup=cat_kb(uid, s))
    return CHOOSE_CATEGORY

async def add_category(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    key = cat_key_from_label(text, s["language"], uid)
    if not key:
        await update.message.reply_text(tr(uid, "choose_cat", s), reply_markup=cat_kb(uid, s))
        return CHOOSE_CATEGORY
    ctx.user_data["category"] = key
    await update.message.reply_text(
        tr(uid, "enter_amount", s, cat=cat_label(s["language"], key, uid), sym=sym(s)),
        parse_mode="HTML", reply_markup=cancel_kb(uid, s)
    )
    return ENTER_AMOUNT

async def add_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    try:
        amount = float(text.replace(",", "."))
        if amount <= 0 or amount > 999_999.99:
            raise ValueError
    except ValueError:
        await update.message.reply_text(tr(uid, "bad_amount", s))
        return ENTER_AMOUNT
    ctx.user_data["amount"] = amount
    await update.message.reply_text(tr(uid, "enter_name", s), reply_markup=cancel_kb(uid, s))
    return ENTER_NAME

async def add_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    amount, cat_key, name = ctx.user_data["amount"], ctx.user_data["category"], text.strip()
    with get_db() as db:
        db.execute("INSERT INTO expenses (user_id,amount,category,item_name) VALUES (?,?,?,?)",
                   (uid, amount, cat_key, name))
    warning = check_limit(uid, cat_key, s)
    msg = tr(uid, "saved", s, name=name, amount=amount,
             cat=cat_label(s["language"], cat_key, uid), sym=sym(s))
    if warning:
        msg += f"\n\n{warning}"
    await update.message.reply_text(msg, parse_mode="HTML", reply_markup=main_kb(uid, s))
    return ConversationHandler.END

# ── Шаблони ────────────────────────────────────────────────────────────────────
def get_templates(uid: int) -> list:
    with get_db() as db:
        return [dict(r) for r in db.execute(
            "SELECT * FROM templates WHERE user_id=? ORDER BY name", (uid,)).fetchall()]

async def tmpl_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    templates = get_templates(uid)
    ctx.user_data["tmpl_map"] = {
        f"⚡ {t['name']} ({t['amount']:.0f} {sym(s)})": t for t in templates
    }
    if not templates:
        await update.message.reply_text(tr(uid, "tmpl_none", s),
            reply_markup=ReplyKeyboardMarkup(
                [[tr(uid, "btn_tmpl_add", s)], [tr(uid, "btn_cancel", s)]],
                resize_keyboard=True))
    else:
        await update.message.reply_text(tr(uid, "tmpl_choose", s),
                                        reply_markup=tmpl_kb(uid, s, templates))
    return TMPL_ACTION

async def tmpl_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    if text == tr(uid, "btn_tmpl_add", s):
        await update.message.reply_text(tr(uid, "tmpl_add_name", s), reply_markup=cancel_kb(uid, s))
        return TMPL_ADD_NAME
    if text == tr(uid, "btn_tmpl_del", s):
        templates = get_templates(uid)
        if not templates:
            await update.message.reply_text(tr(uid, "tmpl_no_del", s), reply_markup=main_kb(uid, s))
            return ConversationHandler.END
        ctx.user_data["tmpl_del_map"] = {t["name"]: t["id"] for t in templates}
        btns = [[t["name"]] for t in templates] + [[tr(uid, "btn_cancel", s)]]
        await update.message.reply_text(tr(uid, "tmpl_del_choose", s),
            reply_markup=ReplyKeyboardMarkup(btns, resize_keyboard=True, one_time_keyboard=True))
        return TMPL_DEL
    tmpl = ctx.user_data.get("tmpl_map", {}).get(text)
    if tmpl:
        with get_db() as db:
            db.execute("INSERT INTO expenses (user_id,amount,category,item_name) VALUES (?,?,?,?)",
                       (uid, tmpl["amount"], tmpl["category"], tmpl["name"]))
        warning = check_limit(uid, tmpl["category"], s)
        msg = tr(uid, "saved", s, name=tmpl["name"], amount=tmpl["amount"],
                 cat=cat_label(s["language"], tmpl["category"], uid), sym=sym(s))
        if warning:
            msg += f"\n\n{warning}"
        await update.message.reply_text(msg, parse_mode="HTML", reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    return TMPL_ACTION

async def tmpl_add_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    ctx.user_data["tmpl_name"] = text.strip()
    await update.message.reply_text(
        tr(uid, "tmpl_add_amount", s, name=ctx.user_data["tmpl_name"], sym=sym(s)),
        reply_markup=cancel_kb(uid, s))
    return TMPL_ADD_AMOUNT

async def tmpl_add_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    try:
        amount = float(text.replace(",", "."))
        if amount <= 0 or amount > 999_999.99:
            raise ValueError
    except ValueError:
        await update.message.reply_text(tr(uid, "bad_amount", s))
        return TMPL_ADD_AMOUNT
    ctx.user_data["tmpl_amount"] = amount
    await update.message.reply_text(
        tr(uid, "tmpl_add_cat", s, name=ctx.user_data["tmpl_name"]),
        reply_markup=cat_kb(uid, s))
    return TMPL_ADD_CAT

async def tmpl_add_cat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    key = cat_key_from_label(text, s["language"], uid)
    if not key:
        return TMPL_ADD_CAT
    name, amount = ctx.user_data["tmpl_name"], ctx.user_data["tmpl_amount"]
    with get_db() as db:
        db.execute("INSERT INTO templates (user_id,name,amount,category) VALUES (?,?,?,?)",
                   (uid, name, amount, key))
    await update.message.reply_text(
        tr(uid, "tmpl_added", s, name=name, amount=amount,
           cat=cat_label(s["language"], key, uid), sym=sym(s)),
        reply_markup=main_kb(uid, s))
    return ConversationHandler.END

async def tmpl_del(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    tmpl_id = ctx.user_data.get("tmpl_del_map", {}).get(text)
    if tmpl_id:
        with get_db() as db:
            db.execute("DELETE FROM templates WHERE id=?", (tmpl_id,))
        await update.message.reply_text(tr(uid, "tmpl_deleted", s, name=text),
                                        reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    return TMPL_DEL

# ── Звіти ──────────────────────────────────────────────────────────────────────
async def summary(update: Update, ctx: ContextTypes.DEFAULT_TYPE, period: str):
    uid = update.effective_user.id
    s = get_settings(uid)
    start, end = period_dates(period)
    rates = await get_rates()
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM expenses WHERE user_id=? AND created BETWEEN ? AND ? ORDER BY created",
            (uid, start, end)
        ).fetchall()
    text = fmt_today(uid, rows, s, rates) if period == "day" else \
           fmt_period(uid, rows, "title_week" if period == "week" else "title_month", s, rates)
    await update.message.reply_text(text, parse_mode="HTML", reply_markup=main_kb(uid, s))

async def summary_day(u, c):   await summary(u, c, "day")
async def summary_week(u, c):  await summary(u, c, "week")
async def summary_month(u, c): await summary(u, c, "month")

async def top_categories(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    start, end = period_dates("month")
    rates = await get_rates()
    _sym = sym(s)
    with get_db() as db:
        rows = db.execute(
            "SELECT category, SUM(amount) as total FROM expenses "
            "WHERE user_id=? AND created BETWEEN ? AND ? GROUP BY category ORDER BY total DESC LIMIT 10",
            (uid, start, end)
        ).fetchall()
    if not rows:
        await update.message.reply_text(tr(uid, "no_data", s), reply_markup=main_kb(uid, s))
        return
    grand = sum(r["total"] for r in rows)
    sec = secondary_str(grand, s, rates)
    lines = [tr(uid, "title_top_cat", s),
             f"{tr(uid, 'total_label', s)}: <b>{grand:.2f} {_sym}{sec}</b>", ""]
    for i, r in enumerate(rows, 1):
        pct = r["total"] / grand * 100
        lines.append(f"{i}. {cat_label(s['language'], r['category'], uid)}: "
                     f"<b>{r['total']:.2f} {_sym}</b>  ({pct:.0f}%)")
    await update.message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=main_kb(uid, s))

async def top_items(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    start, end = period_dates("month")
    _sym = sym(s)
    with get_db() as db:
        rows = db.execute(
            "SELECT item_name, category, SUM(amount) as total, COUNT(*) as cnt "
            "FROM expenses WHERE user_id=? AND created BETWEEN ? AND ? "
            "GROUP BY item_name, category ORDER BY total DESC LIMIT 10",
            (uid, start, end)
        ).fetchall()
    if not rows:
        await update.message.reply_text(tr(uid, "no_data", s), reply_markup=main_kb(uid, s))
        return
    lines = [tr(uid, "title_top_items", s), ""]
    for i, r in enumerate(rows, 1):
        lines.append(f"{i}. <b>{r['item_name']}</b> [{cat_label(s['language'], r['category'], uid)}]\n"
                     f"   {r['total']:.2f} {_sym} {tr(uid, 'times', s)} {r['cnt']}")
    await update.message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=main_kb(uid, s))

# ── Ліміт ──────────────────────────────────────────────────────────────────────
async def limit_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    await update.message.reply_text(tr(uid, "limit_choose", s), reply_markup=cat_kb(uid, s))
    return SET_LIMIT_CAT

async def limit_category(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    key = cat_key_from_label(text, s["language"], uid)
    if not key:
        await update.message.reply_text(tr(uid, "limit_choose", s), reply_markup=cat_kb(uid, s))
        return SET_LIMIT_CAT
    ctx.user_data["limit_cat"] = key
    await update.message.reply_text(
        tr(uid, "limit_enter", s, cat=cat_label(s["language"], key, uid), sym=sym(s)),
        reply_markup=cancel_kb(uid, s))
    return SET_LIMIT_AMOUNT

async def limit_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    try:
        amount = float(text.replace(",", "."))
        if amount < 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(tr(uid, "bad_amount", s))
        return SET_LIMIT_AMOUNT
    cat_key = ctx.user_data["limit_cat"]
    cat = cat_label(s["language"], cat_key, uid)
    with get_db() as db:
        if amount == 0:
            db.execute("DELETE FROM limits WHERE user_id=? AND category=?", (uid, cat_key))
            msg = tr(uid, "limit_removed", s, cat=cat)
        else:
            db.execute("INSERT INTO limits (user_id,category,amount) VALUES (%s,%s,%s) "
            "ON CONFLICT (user_id,category) DO UPDATE SET amount=EXCLUDED.amount",
                       (uid, cat_key, amount))
            msg = tr(uid, "limit_set", s, cat=cat, amount=amount, sym=sym(s))
    await update.message.reply_text(msg, parse_mode="HTML", reply_markup=main_kb(uid, s))
    return ConversationHandler.END

# ── Видалити останню ───────────────────────────────────────────────────────────
async def delete_last(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s = get_settings(uid)
    with get_db() as db:
        row = db.execute("SELECT * FROM expenses WHERE user_id=? ORDER BY id DESC LIMIT 1",
                         (uid,)).fetchone()
        if not row:
            await update.message.reply_text(tr(uid, "nothing_del", s), reply_markup=main_kb(uid, s))
            return
        db.execute("DELETE FROM expenses WHERE id=?", (row["id"],))
    await update.message.reply_text(
        tr(uid, "deleted", s, name=row["item_name"], amount=row["amount"],
           cat=cat_label(s["language"], row["category"], uid), sym=sym(s)),
        parse_mode="HTML", reply_markup=main_kb(uid, s))



# ── Онбординг ─────────────────────────────────────────────────────────────────
async def onboard_lang(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = update.message.text.strip()
    lang_map = {v: k for k, v in LANG_BUTTONS.items()}
    if text in lang_map:
        save_settings(uid, language=lang_map[text])
        s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "onboard_cur_pri", s), parse_mode="HTML",
        reply_markup=curr_kb(uid, s, with_none=False)
    )
    return ONBOARD_CUR_PRI

async def onboard_cur_pri(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = update.message.text.strip()
    cur_map = {v: k for k, v in CURRENCY_BUTTONS.items()}
    if text in cur_map:
        save_settings(uid, primary_currency=cur_map[text])
        s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "onboard_cur_sec", s), parse_mode="HTML",
        reply_markup=curr_kb(uid, s, with_none=True)
    )
    return ONBOARD_CUR_SEC

async def onboard_cur_sec(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = update.message.text.strip()
    cur_map = {v: k for k, v in CURRENCY_BUTTONS.items()}
    if text in cur_map:
        save_settings(uid, secondary_currency=cur_map[text])
    elif text == tr(uid, "curr_none", s):
        save_settings(uid, secondary_currency=None)
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "onboard_done", s), parse_mode="HTML",
        reply_markup=main_kb(uid, s)
    )
    return ConversationHandler.END


# ── Налаштування (Settings menu) ─────────────────────────────────────────────
async def settings_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "settings_title", s), parse_mode="HTML",
        reply_markup=settings_kb(uid, s)
    )
    return SETTINGS_MENU

async def settings_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = update.message.text.strip()

    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END

    if text == tr(uid, "btn_lang", s):
        await update.message.reply_text(tr(uid, "lang_select", s), reply_markup=lang_kb(uid, s))
        return LANG_SELECT

    if text == tr(uid, "btn_currency", s):
        await update.message.reply_text(
            tr(uid, "curr_primary", s), parse_mode="HTML",
            reply_markup=curr_kb(uid, s, with_none=False)
        )
        return CURR_PRIMARY

    if text == tr(uid, "btn_donate", s):
        if not DONATE_URL:
            await update.message.reply_text(tr(uid, "donate_no_url", s), reply_markup=main_kb(uid, s))
        else:
            await update.message.reply_text(
                tr(uid, "donate_msg", s, url=DONATE_URL),
                parse_mode="HTML", reply_markup=main_kb(uid, s),
                disable_web_page_preview=True
            )
        return ConversationHandler.END

    await update.message.reply_text(
        tr(uid, "settings_title", s), parse_mode="HTML",
        reply_markup=settings_kb(uid, s)
    )
    return SETTINGS_MENU


# ── Дохід ─────────────────────────────────────────────────────────────────────
async def income_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "income_enter_amount", s, sym=sym(s)),
        parse_mode="HTML", reply_markup=cancel_kb(uid, s)
    )
    return INCOME_AMT

async def income_got_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    s   = get_settings(uid)
    text = sanitize(update.message.text)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    amount = parse_amount(text)
    if not amount:
        await update.message.reply_text(tr(uid, "bad_amount", s), reply_markup=cancel_kb(uid, s))
        return INCOME_AMT
    ctx.user_data["inc_amt"] = amount
    await update.message.reply_text(tr(uid, "income_enter_source", s), reply_markup=cancel_kb(uid, s))
    return INCOME_SRC

async def income_got_source(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid    = update.effective_user.id
    s      = get_settings(uid)
    source = sanitize(update.message.text, 60)
    if source == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    amount = ctx.user_data["inc_amt"]
    add_income(uid, amount, source)
    after = ctx.user_data.pop("after_income", None)
    kb = finance_kb(uid, s) if after == "finance" else main_kb(uid, s)
    await update.message.reply_text(
        tr(uid, "income_saved", s, source=source, amount=amount, sym=sym(s)),
        parse_mode="HTML", reply_markup=kb
    )
    return ConversationHandler.END


# ── Баланс ────────────────────────────────────────────────────────────────────
async def show_balance(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return
    s    = get_settings(uid)
    lang = s["language"]
    _sym = sym(s)
    now  = datetime.now()

    MONTHS_SHORT = {
        "uk":["","Січ","Лют","Бер","Кві","Тра","Чер","Лип","Сер","Вер","Жов","Лис","Гру"],
        "ru":["","Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"],
        "en":["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],
        "de":["","Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"],
    }
    month_name = f"{MONTHS_SHORT[lang][now.month]} {now.year}"

    income_total   = get_month_income(uid)
    expenses_total = get_month_expenses_total(uid)

    if income_total == 0 and expenses_total == 0:
        await update.message.reply_text(tr(uid, "balance_no_data", s), reply_markup=main_kb(uid, s))
        return

    balance = income_total - expenses_total
    lines = [
        tr(uid, "balance_title",    s, month=month_name),
        tr(uid, "balance_income",   s, amount=income_total,   sym=_sym),
        tr(uid, "balance_expenses", s, amount=expenses_total, sym=_sym),
        "",
        tr(uid, "balance_result_pos" if balance >= 0 else "balance_result_neg",
           s, amount=abs(balance), sym=_sym),
    ]
    await update.message.reply_text(
        "\n".join(lines), parse_mode="HTML", reply_markup=main_kb(uid, s)
    )


# ── Інструкція ────────────────────────────────────────────────────────────────
async def show_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "help_text", s), parse_mode="HTML",
        reply_markup=main_kb(uid, s),
        disable_web_page_preview=True
    )


# ── Фінанси (підменю) ─────────────────────────────────────────────────────────
async def finance_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s    = get_settings(uid)
    lang = s["language"]
    _sym = sym(s)
    now  = datetime.now()

    MONTHS_SHORT = {
        "uk":["","Січ","Лют","Бер","Кві","Тра","Чер","Лип","Сер","Вер","Жов","Лис","Гру"],
        "ru":["","Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"],
        "en":["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],
        "de":["","Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"],
    }
    month_name = f"{MONTHS_SHORT[lang][now.month]} {now.year}"

    income_total   = get_month_income(uid)
    expenses_total = get_month_expenses_total(uid)
    balance        = income_total - expenses_total

    lines = [tr(uid, "finance_title", s, month=month_name), ""]
    if income_total > 0:
        lines.append(tr(uid, "balance_income",   s, amount=income_total,   sym=_sym))
    else:
        lines.append(tr(uid, "finance_no_income", s))
    lines.append(tr(uid, "balance_expenses", s, amount=expenses_total, sym=_sym))
    if income_total > 0:
        lines.append("")
        lines.append(tr(uid, "balance_result_pos" if balance >= 0 else "balance_result_neg",
                       s, amount=abs(balance), sym=_sym))
    lines.append("")
    lines.append(tr(uid, "finance_add_income_hint", s))

    await update.message.reply_text(
        "\n".join(lines), parse_mode="HTML",
        reply_markup=finance_kb(uid, s)
    )
    return FINANCE_MENU

async def finance_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = sanitize(update.message.text)
    if any(text == T[l].get("btn_back","") for l in T):
        await update.message.reply_text(tr(uid, "choose_menu", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    if any(text == T[l].get("btn_add_income_short","") for l in T):
        await update.message.reply_text(
            tr(uid, "income_enter_amount", s, sym=sym(s)),
            parse_mode="HTML", reply_markup=cancel_kb(uid, s)
        )
        ctx.user_data["after_income"] = "finance"
        return INCOME_AMT
    return FINANCE_MENU


# ── Звіти (підменю) ───────────────────────────────────────────────────────────
async def reports_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "choose_menu", s), reply_markup=reports_kb(uid, s)
    )
    return REPORTS_MENU

async def reports_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = sanitize(update.message.text)

    if any(text == T[l].get("btn_back","") for l in T):
        await update.message.reply_text(tr(uid, "choose_menu", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END

    # Route to existing handlers — show reports_kb after each report
    handled = False
    for lang in T:
        if text == T[lang].get("btn_today"):   await summary_day(update, ctx);    handled = True; break
        if text == T[lang].get("btn_week"):    await summary_week(update, ctx);   handled = True; break
        if text == T[lang].get("btn_month"):   await summary_month(update, ctx);  handled = True; break
        if text == T[lang].get("btn_compare"): await compare_months(update, ctx); handled = True; break
        if text == T[lang].get("btn_top_cat"): await top_categories(update, ctx); handled = True; break
        if text == T[lang].get("btn_top_items"):await top_items(update, ctx);     handled = True; break

    if handled:
        # Re-show reports submenu after displaying the report
        await update.message.reply_text(tr(uid, "choose_menu", s), reply_markup=reports_kb(uid, s))
        return REPORTS_MENU

    # Excel
    if any(text == T[l].get("btn_export","") for l in T):
        await export_start(update, ctx)
        return ConversationHandler.END

    return REPORTS_MENU


# ── Більше (підменю) ──────────────────────────────────────────────────────────
async def more_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "choose_menu", s), reply_markup=more_kb(uid, s)
    )
    return MORE_MENU

async def more_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = sanitize(update.message.text)

    if any(text == T[l].get("btn_back","") for l in T):
        await update.message.reply_text(tr(uid, "choose_menu", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    if text == tr(uid, "btn_help", s):
        await show_help(update, ctx); return MORE_MENU
    if text == tr(uid, "btn_feedback", s):
        await feedback_start(update, ctx); return ConversationHandler.END
    if text == tr(uid, "btn_convert", s):
        await convert_start(update, ctx); return ConversationHandler.END
    if text == tr(uid, "btn_limit", s):
        await limit_start(update, ctx); return ConversationHandler.END
    if text == tr(uid, "btn_recurring", s):
        await recur_start(update, ctx); return ConversationHandler.END
    if text == tr(uid, "btn_reminders", s):
        await remind_start(update, ctx); return ConversationHandler.END
    if text == tr(uid, "btn_my_cats", s):
        await ccat_start(update, ctx); return ConversationHandler.END

    return MORE_MENU

# ── Порівняння місяців ────────────────────────────────────────────────────────
async def compare_months(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return
    s = get_settings(uid)
    lang = s["language"]
    _sym = sym(s)
    fmt = "%Y-%m-%d %H:%M:%S"
    now = datetime.now()
    cur_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_end   = cur_start - timedelta(seconds=1)
    prev_start = prev_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    def _rows(start, end):
        with get_db() as db:
            return db.execute(
                "SELECT * FROM expenses WHERE user_id=? AND created BETWEEN ? AND ?",
                (uid, start.strftime(fmt), end.strftime(fmt))
            ).fetchall()

    cur_rows  = _rows(cur_start, now)
    prev_rows = _rows(prev_start, prev_end)
    cur_total  = sum(float(r["amount"]) for r in cur_rows)
    prev_total = sum(float(r["amount"]) for r in prev_rows)

    if not cur_rows and not prev_rows:
        await update.message.reply_text(tr(uid,"compare_no_data",s), reply_markup=main_kb(uid,s))
        return

    MONTHS = {
        "uk":["","січ","лют","бер","кві","тра","чер","лип","сер","вер","жов","лис","гру"],
        "ru":["","янв","фев","мар","апр","май","июн","июл","авг","сен","окт","ноя","дек"],
        "en":["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],
        "de":["","Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"],
    }
    cur_m  = MONTHS[lang][now.month]
    prev_m = MONTHS[lang][prev_end.month]

    lines = [tr(uid,"compare_title",s), ""]
    lines.append(f"<b>{tr(uid,'compare_this',s)} ({cur_m})</b>: {cur_total:.2f} {_sym}")
    lines.append(f"<b>{tr(uid,'compare_prev',s)} ({prev_m})</b>: {prev_total:.2f} {_sym}")
    lines.append("")

    if prev_total > 0:
        diff = cur_total - prev_total; pct = abs(diff)/prev_total*100
        if diff > 0:   lines.append(tr(uid,"compare_diff_more",s,diff=diff,pct=pct,sym=_sym))
        elif diff < 0: lines.append(tr(uid,"compare_diff_less",s,diff=abs(diff),pct=pct,sym=_sym))
        else:          lines.append(tr(uid,"compare_diff_same",s))
    elif cur_total > 0:
        lines.append(tr(uid,"compare_diff_more",s,diff=cur_total,pct=100,sym=_sym))

    from collections import defaultdict
    all_cats: set = set()
    cur_cat: dict[str,float] = defaultdict(float)
    prev_cat: dict[str,float] = defaultdict(float)
    for r in cur_rows:  cur_cat[r["category"]] += float(r["amount"]); all_cats.add(r["category"])
    for r in prev_rows: prev_cat[r["category"]] += float(r["amount"]); all_cats.add(r["category"])
    if all_cats:
        lines.append(""); lines.append(f"<b>{tr(uid,'compare_by_cat',s)}</b>")
        for cat in sorted(all_cats, key=lambda c: -(cur_cat.get(c,0)+prev_cat.get(c,0))):
            cl = cat_label(lang, cat, uid)
            c_v, p_v = cur_cat.get(cat,0), prev_cat.get(cat,0)
            arrow = "▲" if c_v>p_v else ("▼" if c_v<p_v else "◆")
            lines.append(f"  {arrow} {cl}: {c_v:.2f} vs {p_v:.2f} {_sym}")

    await update.message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=main_kb(uid,s))


# ── Регулярні витрати ─────────────────────────────────────────────────────────
async def recur_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    items = get_recurring(uid)
    await update.message.reply_text(
        tr(uid,"recur_choose",s) if items else tr(uid,"recur_none",s),
        reply_markup=recur_kb(uid, s, items)
    )
    return RECUR_MENU

async def recur_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text)
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s))
        return ConversationHandler.END
    if text == tr(uid,"btn_recur_add",s):
        await update.message.reply_text(tr(uid,"recur_add_name",s), reply_markup=cancel_kb(uid,s))
        return RECUR_NAME
    if text == tr(uid,"btn_recur_del",s):
        items = get_recurring(uid)
        if not items:
            await update.message.reply_text(tr(uid,"recur_no_del",s), reply_markup=main_kb(uid,s))
            return ConversationHandler.END
        await update.message.reply_text(tr(uid,"recur_del_choose",s), reply_markup=recur_del_kb(uid,s,items))
        return RECUR_DEL
    items = get_recurring(uid)
    await update.message.reply_text(tr(uid,"recur_choose",s), reply_markup=recur_kb(uid,s,items))
    return RECUR_MENU

async def recur_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    name = sanitize(update.message.text, 50)
    if name == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    ctx.user_data["rc_name"] = name
    await update.message.reply_text(tr(uid,"recur_add_amount",s,name=name,sym=sym(s)), reply_markup=cancel_kb(uid,s))
    return RECUR_AMT

async def recur_amt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    amount = parse_amount(update.message.text)
    if not amount:
        await update.message.reply_text(tr(uid,"bad_amount",s), reply_markup=cancel_kb(uid,s)); return RECUR_AMT
    ctx.user_data["rc_amt"] = amount
    await update.message.reply_text(tr(uid,"recur_add_cat",s,name=ctx.user_data["rc_name"]), reply_markup=cat_kb(uid,s))
    return RECUR_CAT

async def recur_cat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    label = sanitize(update.message.text)
    if label == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    key = cat_key_from_label(label, s["language"], uid)
    if not key:
        await update.message.reply_text(tr(uid,"choose_cat",s), reply_markup=cat_kb(uid,s)); return RECUR_CAT
    ctx.user_data["rc_cat"] = key
    await update.message.reply_text(tr(uid,"recur_add_day",s), reply_markup=cancel_kb(uid,s))
    return RECUR_DAY

async def recur_day(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text)
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    try:
        day = int(text)
        if not 1 <= day <= 31: raise ValueError
    except ValueError:
        await update.message.reply_text(tr(uid,"recur_bad_day",s), reply_markup=cancel_kb(uid,s)); return RECUR_DAY
    name   = ctx.user_data["rc_name"]
    amount = ctx.user_data["rc_amt"]
    cat    = ctx.user_data["rc_cat"]
    add_recurring_db(uid, name, amount, cat, day)
    await update.message.reply_text(
        tr(uid,"recur_added",s,name=name,day=day,amount=amount,sym=sym(s)),
        reply_markup=main_kb(uid,s)
    )
    return ConversationHandler.END

async def recur_del(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text)
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    items = get_recurring(uid)
    for item in items:
        if text == f"🗑 {item['name']}":
            del_recurring_db(item["id"])
            await update.message.reply_text(tr(uid,"recur_deleted",s,name=item["name"]), reply_markup=main_kb(uid,s))
            return ConversationHandler.END
    await update.message.reply_text(tr(uid,"recur_del_choose",s), reply_markup=recur_del_kb(uid,s,items))
    return RECUR_DEL


# ── Нагадування ──────────────────────────────────────────────────────────────
def _valid_hhmm(t: str) -> bool:
    parts = t.split(":")
    if len(parts) != 2: return False
    try: h,m = int(parts[0]),int(parts[1]); return 0<=h<=23 and 0<=m<=59
    except ValueError: return False

def _fmt_remind_status(uid: int, s: dict) -> str:
    rs = get_reminder_settings(uid)
    off = tr(uid,"remind_off",s)
    inact  = str(rs["inactive_days"]) if rs["inactive_days"] else off
    daily  = rs["daily_time"]  or off
    weekly = rs["weekly_time"] or off
    wday   = str(rs["weekly_day"]) if rs["weekly_day"] is not None else off
    return tr(uid,"remind_current",s,inact=inact,daily=daily,weekly=weekly,wday=wday)

async def remind_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid,"remind_title",s) + "\n\n" + _fmt_remind_status(uid,s),
        parse_mode="HTML", reply_markup=remind_kb(uid,s)
    )
    return REMIND_MENU

async def remind_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text)
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    if text == tr(uid,"remind_btn_inactive",s):
        await update.message.reply_text(tr(uid,"remind_inactive_prompt",s), reply_markup=cancel_kb(uid,s)); return REMIND_INACT
    if text == tr(uid,"remind_btn_daily",s):
        await update.message.reply_text(tr(uid,"remind_daily_prompt",s), reply_markup=cancel_kb(uid,s)); return REMIND_DAILY
    if text == tr(uid,"remind_btn_weekly",s):
        await update.message.reply_text(tr(uid,"remind_weekly_day_prompt",s), reply_markup=cancel_kb(uid,s)); return REMIND_WDAY
    await update.message.reply_text(
        tr(uid,"remind_title",s)+"\n\n"+_fmt_remind_status(uid,s),
        parse_mode="HTML", reply_markup=remind_kb(uid,s)
    )
    return REMIND_MENU

async def remind_set_inact(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text)
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    try:
        days = int(text)
        if not 0 <= days <= 30: raise ValueError
    except ValueError:
        await update.message.reply_text(tr(uid,"remind_bad_inactive",s), reply_markup=cancel_kb(uid,s)); return REMIND_INACT
    save_reminder_settings(uid, inactive_days=days)
    await update.message.reply_text(tr(uid,"remind_saved",s), reply_markup=main_kb(uid,s))
    return ConversationHandler.END

async def remind_set_daily(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text).strip()
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    if text == "0":
        save_reminder_settings(uid, daily_time=None)
        await update.message.reply_text(tr(uid,"remind_saved",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    if not _valid_hhmm(text):
        await update.message.reply_text(tr(uid,"remind_bad_time",s), reply_markup=cancel_kb(uid,s)); return REMIND_DAILY
    h,m = text.split(":")
    save_reminder_settings(uid, daily_time=f"{int(h):02d}:{int(m):02d}")
    await update.message.reply_text(tr(uid,"remind_saved",s), reply_markup=main_kb(uid,s))
    return ConversationHandler.END

async def remind_set_wday(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text).strip()
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    try:
        day = int(text)
        if not -1 <= day <= 6: raise ValueError
    except ValueError:
        await update.message.reply_text(tr(uid,"remind_bad_wday",s), reply_markup=cancel_kb(uid,s)); return REMIND_WDAY
    if day == -1:
        save_reminder_settings(uid, weekly_day=None, weekly_time=None)
        await update.message.reply_text(tr(uid,"remind_saved",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    ctx.user_data["rm_wday"] = day
    await update.message.reply_text(tr(uid,"remind_weekly_time_prompt",s), reply_markup=cancel_kb(uid,s))
    return REMIND_WTIME

async def remind_set_wtime(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id; s = get_settings(uid)
    text = sanitize(update.message.text).strip()
    if text == tr(uid,"btn_cancel",s):
        await update.message.reply_text(tr(uid,"cancelled",s), reply_markup=main_kb(uid,s)); return ConversationHandler.END
    if not _valid_hhmm(text):
        await update.message.reply_text(tr(uid,"remind_bad_time",s), reply_markup=cancel_kb(uid,s)); return REMIND_WTIME
    h,m = text.split(":")
    save_reminder_settings(uid, weekly_day=ctx.user_data["rm_wday"],
                           weekly_time=f"{int(h):02d}:{int(m):02d}")
    await update.message.reply_text(tr(uid,"remind_saved",s), reply_markup=main_kb(uid,s))
    return ConversationHandler.END


# ── Планувальник ──────────────────────────────────────────────────────────────
async def job_recurring(context):
    """Daily at midnight UTC — auto-insert recurring expenses."""
    import datetime as _dt, calendar
    today      = _dt.date.today()
    today_day  = today.day
    last_day   = calendar.monthrange(today.year, today.month)[1]
    for rec in get_all_recurring():
        # Fire on configured day, or on last day of month if month is shorter
        effective_day = min(rec["day_of_month"], last_day)
        if effective_day != today_day:
            continue
        uid = rec["user_id"]; s = get_settings(uid)
        with get_db() as db:
            db.execute(
                "INSERT INTO expenses (user_id,amount,category,item_name) VALUES (?,?,?,?)",
                (uid, rec["amount"], rec["category"], rec["name"])
            )
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=tr(uid,"recur_auto",s,name=rec["name"],amount=rec["amount"],sym=sym(s)),
                parse_mode="HTML"
            )
        except Exception:
            pass

async def job_reminders(context):
    """Every 5 min — check per-user reminder schedules (UTC)."""
    import datetime as _dt
    now = _dt.datetime.utcnow()
    current_hhmm    = now.strftime("%H:%M")
    current_weekday = now.weekday()

    for rs in get_all_users_with_reminders():
        uid = rs["user_id"]; s = get_settings(uid)
        if rs["inactive_days"] > 0 and current_hhmm == "09:00":
            last = get_last_expense_date(uid)
            if last:
                last_date = last.date() if hasattr(last, "date") else _dt.datetime.strptime(str(last)[:10], "%Y-%m-%d").date()
                inactive  = (_dt.date.today() - last_date).days
                if inactive >= rs["inactive_days"]:
                    try:
                        await context.bot.send_message(
                            chat_id=uid, parse_mode="HTML",
                            text=tr(uid,"remind_inact_msg",s,days=inactive)
                        )
                    except Exception: pass
        if rs.get("daily_time") and rs["daily_time"] == current_hhmm:
            try:
                await context.bot.send_message(
                    chat_id=uid, parse_mode="HTML",
                    text=tr(uid,"remind_daily_msg",s,btn=tr(uid,"btn_today",s))
                )
            except Exception: pass
        if (rs.get("weekly_day") is not None and rs.get("weekly_time")
                and rs["weekly_day"] == current_weekday
                and rs["weekly_time"] == current_hhmm):
            try:
                await context.bot.send_message(
                    chat_id=uid, parse_mode="HTML",
                    text=tr(uid,"remind_weekly_msg",s,btn=tr(uid,"btn_week",s))
                )
            except Exception: pass


# ── Зворотній зв'язок ─────────────────────────────────────────────────────────
async def feedback_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not _is_allowed(uid): return ConversationHandler.END
    s = get_settings(uid)
    await update.message.reply_text(
        tr(uid, "feedback_prompt", s), reply_markup=cancel_kb(uid, s)
    )
    return FEEDBACK_MSG

async def feedback_got_msg(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    s    = get_settings(uid)
    text = sanitize(update.message.text, 1000)
    if text == tr(uid, "btn_cancel", s):
        await update.message.reply_text(tr(uid, "cancelled", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END

    # Rate limit: 1 feedback per hour per user
    now = time.time()
    if now - _feedback_cooldown.get(uid, 0) < 3600:
        await update.message.reply_text(tr(uid, "feedback_cooldown", s), reply_markup=main_kb(uid, s))
        return ConversationHandler.END
    _feedback_cooldown[uid] = now

    user = update.effective_user
    name = (user.full_name or "").strip() or f"User {uid}"
    username = f"@{user.username}" if user.username else "немає username"

    if ADMIN_ID:
        try:
            # Send to admin with reply instructions
            admin_text = (
                f"{tr(uid, 'feedback_received', s, name=name, user_id=uid, text=text)}\n"
                f"Username: {username}\n\n"
                f"<i>Щоб відповісти: /reply {uid} текст відповіді</i>"
            )
            await ctx.bot.send_message(
                chat_id=ADMIN_ID,
                text=admin_text,
                parse_mode="HTML"
            )
        except Exception as e:
            logging.warning("Could not send feedback to admin: %s", e)

    await update.message.reply_text(tr(uid, "feedback_sent", s), reply_markup=main_kb(uid, s))
    return ConversationHandler.END


async def cmd_reply(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin command: /reply USER_ID text — sends reply to user from bot."""
    sender_id = update.effective_user.id
    if sender_id != ADMIN_ID:
        return  # Only admin can use this

    # Parse: /reply 123456 текст
    args = update.message.text.split(maxsplit=2)
    if len(args) < 3:
        s = get_settings(sender_id)
        await update.message.reply_text(tr(sender_id, "feedback_reply_usage", s))
        return

    try:
        target_uid = int(args[1])
    except ValueError:
        s = get_settings(sender_id)
        await update.message.reply_text(tr(sender_id, "feedback_reply_usage", s))
        return

    reply_text = args[2]
    s_target = get_settings(target_uid)
    s_admin  = get_settings(sender_id)

    try:
        await ctx.bot.send_message(
            chat_id=target_uid,
            text=tr(target_uid, "feedback_reply", s_target, text=reply_text),
            parse_mode="HTML"
        )
        await update.message.reply_text(
            tr(sender_id, "feedback_reply_sent", s_admin, uid=target_uid)
        )
    except Exception:
        await update.message.reply_text(
            tr(sender_id, "feedback_reply_fail", s_admin)
        )

# ── Маршрутизатор ──────────────────────────────────────────────────────────────
async def menu_router(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid, text = update.effective_user.id, update.message.text
    s = get_settings(uid)
    for lang in T:
        if text == T[lang].get("btn_reports"):  return await reports_menu(update, ctx)
        if text == T[lang].get("btn_more"):     return await more_menu(update, ctx)
        if text == T[lang].get("btn_finance"):  return await finance_menu(update, ctx)
        if text == T[lang].get("btn_today"):    return await summary_day(update, ctx)
        if text == T[lang].get("btn_week"):     return await summary_week(update, ctx)
        if text == T[lang].get("btn_month"):    return await summary_month(update, ctx)
        if text == T[lang].get("btn_top_cat"):  return await top_categories(update, ctx)
        if text == T[lang].get("btn_top_items"):return await top_items(update, ctx)
        if text == T[lang].get("btn_delete"):   return await delete_last(update, ctx)
        if text == T[lang].get("btn_donate"):   return await cmd_donate(update, ctx)
        if text == T[lang].get("btn_compare"):  return await compare_months(update, ctx)
        if text == T[lang].get("btn_balance"):  return await show_balance(update, ctx)
        if text == T[lang].get("btn_help"):     return await show_help(update, ctx)
        if text == T[lang].get("btn_settings"): return await settings_start(update, ctx)
        if text == T[lang].get("btn_back"):
            await update.message.reply_text(tr(uid, "choose_menu", s), reply_markup=main_kb(uid, s))
            return
        if text == T[lang].get("btn_add"):      return await add_start(update, ctx)
        if text == T[lang].get("btn_quick"):    return await tmpl_start(update, ctx)
    await update.message.reply_text(tr(uid, "choose_menu", s), reply_markup=main_kb(uid, s))

# ── Запуск ─────────────────────────────────────────────────────────────────────
def make_pat(*keys) -> str:
    parts = [re.escape(T[l][k]) for l in T for k in keys if k in T[l]]
    return f"^({'|'.join(parts)})$"

def _is_menu_button(text: str) -> bool:
    """Returns True if text matches any button in any language."""
    return any(text == T[l][k] for l in T for k in T[l])

async def _menu_escape(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Universal fallback — если нажата кнопка меню внутри диалога, выходим."""
    await menu_router(update, ctx)
    return ConversationHandler.END

def main():
    init_db()
    app = ApplicationBuilder().token(TOKEN).build()

    convs = [
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_add")), add_start)],
            states={
                CHOOSE_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_category)],
                ENTER_AMOUNT:    [MessageHandler(filters.TEXT & ~filters.COMMAND, add_amount)],
                ENTER_NAME:      [MessageHandler(filters.TEXT & ~filters.COMMAND, add_name)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_quick")), tmpl_start)],
            states={
                TMPL_ACTION:     [MessageHandler(filters.TEXT & ~filters.COMMAND, tmpl_action)],
                TMPL_ADD_NAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, tmpl_add_name)],
                TMPL_ADD_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, tmpl_add_amount)],
                TMPL_ADD_CAT:    [MessageHandler(filters.TEXT & ~filters.COMMAND, tmpl_add_cat)],
                TMPL_DEL:        [MessageHandler(filters.TEXT & ~filters.COMMAND, tmpl_del)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_limit")), limit_start)],
            states={
                SET_LIMIT_CAT:    [MessageHandler(filters.TEXT & ~filters.COMMAND, limit_category)],
                SET_LIMIT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, limit_amount)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_lang")), lang_start)],
            states={LANG_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, lang_select)]},
            fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_currency")), curr_start)],
            states={
                CURR_PRIMARY:   [MessageHandler(filters.TEXT & ~filters.COMMAND, curr_primary)],
                CURR_SECONDARY: [MessageHandler(filters.TEXT & ~filters.COMMAND, curr_secondary)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_convert")), convert_start)],
            states={CONVERT: [MessageHandler(filters.TEXT & ~filters.COMMAND, convert_do)]},
            fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_export")), export_start)],
            states={EXPORT_PERIOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, export_do)]},
            fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_my_cats")), ccat_start)],
            states={
                CCAT_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, ccat_menu)],
                CCAT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ccat_name)],
                CCAT_DEL:  [MessageHandler(filters.TEXT & ~filters.COMMAND, ccat_del)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_recurring")), recur_start)],
            states={
                RECUR_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, recur_menu)],
                RECUR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, recur_name)],
                RECUR_AMT:  [MessageHandler(filters.TEXT & ~filters.COMMAND, recur_amt)],
                RECUR_CAT:  [MessageHandler(filters.TEXT & ~filters.COMMAND, recur_cat)],
                RECUR_DAY:  [MessageHandler(filters.TEXT & ~filters.COMMAND, recur_day)],
                RECUR_DEL:  [MessageHandler(filters.TEXT & ~filters.COMMAND, recur_del)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_reminders")), remind_start)],
            states={
                REMIND_MENU:  [MessageHandler(filters.TEXT & ~filters.COMMAND, remind_menu)],
                REMIND_INACT: [MessageHandler(filters.TEXT & ~filters.COMMAND, remind_set_inact)],
                REMIND_DAILY: [MessageHandler(filters.TEXT & ~filters.COMMAND, remind_set_daily)],
                REMIND_WDAY:  [MessageHandler(filters.TEXT & ~filters.COMMAND, remind_set_wday)],
                REMIND_WTIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, remind_set_wtime)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_settings")), settings_start)],
            states={
                SETTINGS_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, settings_menu)],
                LANG_SELECT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, lang_select)],
                CURR_PRIMARY:  [MessageHandler(filters.TEXT & ~filters.COMMAND, curr_primary)],
                CURR_SECONDARY:[MessageHandler(filters.TEXT & ~filters.COMMAND, curr_secondary)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_feedback")), feedback_start)],
            states={
                FEEDBACK_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, feedback_got_msg)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_income")), income_start)],
            states={
                INCOME_AMT: [MessageHandler(filters.TEXT & ~filters.COMMAND, income_got_amount)],
                INCOME_SRC: [MessageHandler(filters.TEXT & ~filters.COMMAND, income_got_source)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)]),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_finance")), finance_menu)],
            states={
                FINANCE_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, finance_action)],
                INCOME_AMT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, income_got_amount)],
                INCOME_SRC:   [MessageHandler(filters.TEXT & ~filters.COMMAND, income_got_source)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)],
            allow_reentry=True),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_reports")), reports_menu)],
            states={
                REPORTS_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, reports_action)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)],
            allow_reentry=True),
        ConversationHandler(
            entry_points=[MessageHandler(filters.Regex(make_pat("btn_more")), more_menu)],
            states={
                MORE_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, more_action)],
            }, fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("cancel", cmd_start)],
            allow_reentry=True),
    ]

    # Onboarding — handles /start and first-time setup flow
    app.add_handler(ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            ONBOARD_LANG:    [MessageHandler(filters.TEXT & ~filters.COMMAND, onboard_lang)],
            ONBOARD_CUR_PRI: [MessageHandler(filters.TEXT & ~filters.COMMAND, onboard_cur_pri)],
            ONBOARD_CUR_SEC: [MessageHandler(filters.TEXT & ~filters.COMMAND, onboard_cur_sec)],
        },
        fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, _menu_escape), CommandHandler("start", cmd_start)],
        allow_reentry=True,
    ))
    app.add_handler(CommandHandler("donate", cmd_donate))
    app.add_handler(CommandHandler("reply", cmd_reply))  # Admin: /reply USER_ID text
    for conv in convs:
        app.add_handler(conv)
    app.add_handler(MessageHandler(filters.Regex(make_pat("btn_compare")), compare_months))
    app.add_handler(MessageHandler(filters.Regex(make_pat("btn_balance")), show_balance))
    app.add_handler(MessageHandler(filters.Regex(make_pat("btn_help")),    show_help))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_router))

    # ── Scheduler ─────────────────────────────────────────────────────────────
    import datetime as _dt
    app.job_queue.run_daily(
        job_recurring,
        time=_dt.time(0, 0, 0),
        name="recurring",
    )
    app.job_queue.run_repeating(
        job_reminders,
        interval=300,
        first=30,
        name="reminders",
    )

    print("Bot started. Ctrl+C to stop.")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()