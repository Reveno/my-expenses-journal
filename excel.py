"""
excel.py — builds an Excel workbook with 3 sheets from expense rows.
"""
import io
from collections import defaultdict
from i18n import cat_label
from config import CURRENCY_SYMBOLS

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

_HDR_FILL = "1a4a3a"   # dark emerald — matches bot colour scheme
_ALT_FILL = "e8f5e9"

_HEADERS = {
    "uk": {"date": "Дата", "cat": "Категорія", "name": "Назва",
           "amount": "Сума", "total": "Всього", "day": "День"},
    "ru": {"date": "Дата", "cat": "Категория", "name": "Название",
           "amount": "Сумма", "total": "Итого", "day": "День"},
    "en": {"date": "Date", "cat": "Category",  "name": "Name",
           "amount": "Amount", "total": "Total", "day": "Day"},
    "de": {"date": "Datum", "cat": "Kategorie", "name": "Name",
           "amount": "Betrag", "total": "Gesamt", "day": "Tag"},
}


def _header_row(ws, row: int, cols: list[str]):
    fill = PatternFill("solid", fgColor=_HDR_FILL)
    font = Font(bold=True, color="FFFFFF")
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row, c, val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 45)


def build_xlsx(uid: int, rows: list, s: dict) -> bytes | None:
    if not XLSX_OK:
        return None
    lang = s.get("language", "uk")
    h    = _HEADERS.get(lang, _HEADERS["en"])
    cur  = s.get("primary_currency", "UAH")
    sym  = CURRENCY_SYMBOLS.get(cur, "")
    wb   = openpyxl.Workbook()

    # Sheet 1 — all expenses
    ws1 = wb.active
    ws1.title = h["date"]
    _header_row(ws1, 1, [h["date"], h["cat"], h["name"], f"{h['amount']} ({sym})"])
    total = 0.0
    for i, r in enumerate(rows, 2):
        alt = PatternFill("solid", fgColor=_ALT_FILL if i % 2 == 0 else "FFFFFF")
        amt = float(r["amount"])
        total += amt
        for c, v in enumerate([
            str(r["created"])[:10],
            cat_label(lang, r["category"], uid),
            r["item_name"],
            round(amt, 2),
        ], 1):
            ws1.cell(i, c, v).fill = alt
    tr_row = len(rows) + 2
    ws1.cell(tr_row, 3, h["total"]).font = Font(bold=True)
    ws1.cell(tr_row, 4, round(total, 2)).font = Font(bold=True)
    _auto_width(ws1)

    # Sheet 2 — by category
    ws2 = wb.create_sheet(h["cat"])
    _header_row(ws2, 1, [h["cat"], f"{h['amount']} ({sym})"])
    by_cat: dict[str, float] = defaultdict(float)
    for r in rows:
        by_cat[r["category"]] += float(r["amount"])
    for i, (k, v) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]), 2):
        alt = PatternFill("solid", fgColor=_ALT_FILL if i % 2 == 0 else "FFFFFF")
        ws2.cell(i, 1, cat_label(lang, k, uid)).fill = alt
        ws2.cell(i, 2, round(v, 2)).fill = alt
    ws2.cell(len(by_cat) + 2, 1, h["total"]).font = Font(bold=True)
    ws2.cell(len(by_cat) + 2, 2, round(total, 2)).font = Font(bold=True)
    _auto_width(ws2)

    # Sheet 3 — by day
    ws3 = wb.create_sheet(h["day"])
    _header_row(ws3, 1, [h["day"], f"{h['amount']} ({sym})"])
    by_day: dict[str, float] = defaultdict(float)
    for r in rows:
        by_day[str(r["created"])[:10]] += float(r["amount"])
    for i, (day, v) in enumerate(sorted(by_day.items()), 2):
        alt = PatternFill("solid", fgColor=_ALT_FILL if i % 2 == 0 else "FFFFFF")
        ws3.cell(i, 1, day).fill = alt
        ws3.cell(i, 2, round(v, 2)).fill = alt
    ws3.cell(len(by_day) + 2, 1, h["total"]).font = Font(bold=True)
    ws3.cell(len(by_day) + 2, 2, round(total, 2)).font = Font(bold=True)
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
